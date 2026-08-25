import json
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "indicadores.json"


def _imai():
    payload = json.loads(DATA.read_text(encoding="utf-8"))
    return payload["indicators"]["IMAI"]


def test_imai_schema_18_columnas():
    ind = _imai()
    assert len(ind["columns"]) == 18, f"IMAI debe tener 18 columnas, tiene {len(ind['columns'])}"
    for i, col in enumerate(ind["columns"]):
        assert col["index"] == i


def test_imai_columnas_ordenadas():
    ind = _imai()
    labels = [c["label"] for c in ind["columns"]]
    assert labels[0].lower().startswith("imai") and "índice" in labels[0].lower()
    assert labels[1].lower().startswith("var. mensual")
    assert labels[2].lower().startswith("var. anual")
    assert "acumulado" in labels[5].lower()
    # Componentes: índices (6-9), var. anual (10-13), var. mensual (14-17)
    for i in range(6, 10):
        assert "índice" in labels[i].lower()
    for i in range(10, 14):
        assert "var. anual" in labels[i].lower()
    for i in range(14, 18):
        assert "var. mensual" in labels[i].lower()


def test_imai_variaciones_mensuales_componentes_orden():
    ind = _imai()
    # 14-17: Minería, Energía, Construcción, Manufacturas (mensual)
    names = [ind["columns"][i]["label"] for i in range(14, 18)]
    assert "minería" in names[0].lower()
    assert "energía" in names[1].lower()
    assert "construcción" in names[2].lower()
    assert "manufacturas" in names[3].lower()


def test_imai_datos_recientes_con_mensual():
    ind = _imai()
    obs = ind["observations"][-1]
    assert len(obs["values"]) == 18
    # IMAI total: índice, mensual, anual
    assert obs["values"][0] is not None
    assert obs["values"][1] is not None
    assert obs["values"][2] is not None
    # Componente mensual y anual
    for i in range(14, 18):
        assert obs["values"][i] is not None, f"falta var. mensual col {i}"
    for i in range(10, 14):
        assert obs["values"][i] is not None, f"falta var. anual col {i}"


def test_imai_kpi_resumen_sin_favorable_adverso():
    ind = _imai()
    resumen = ind.get("metrics", {}).get("resumen", [])
    assert resumen, "IMAI debe tener resumen"
    texto = " ".join(resumen).lower()
    assert "favorable" not in texto
    assert "adverso" not in texto
    assert "se clasifica" not in texto


def test_imai_kpi_tiene_acum():
    ind = _imai()
    kpi = ind.get("metrics", {}).get("kpi", {})
    assert kpi.get("acumText") is not None
    assert kpi.get("acumText") != "—"


def test_imai_cifra_actual():
    ind = _imai()
    kpi = ind.get("metrics", {}).get("kpi", {})
    ult = kpi.get("ultimoRaw")
    assert ult is not None
    assert 50 < ult < 130


def test_imai_componentes_ultimo_periodo():
    ind = _imai()
    kpi = ind.get("metrics", {}).get("kpi", {})
    cards = kpi.get("cards", [])
    assert len(cards) == 4
    for c in cards:
        assert "nivelText" in c
        assert "momText" in c
        assert "yoyText" in c


def test_imai_excel_individual_hojas(tmp_path):
    from build_excel import _build_imai_workbook
    ind = _imai()
    out = tmp_path / "IMAI_datos.xlsx"
    _build_imai_workbook(ind, out)
    assert out.exists()

    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert "Niveles" in wb.sheetnames
    assert "Variaciones" in wb.sheetnames
    assert "Resumen" in wb.sheetnames


def test_imai_url_boletin_oficial():
    ind = _imai()
    url = ind.get("url_boletin_oficial")
    assert url
    assert "inegi.org.mx/contenidos/saladeprensa/boletines" in url
    assert url.endswith(".pdf")
