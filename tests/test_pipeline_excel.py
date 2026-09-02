import json
from pathlib import Path

import openpyxl
import pytest

import build_data
import build_excel
import build_calendar
import lib_data as L

ROOT = Path(__file__).resolve().parents[1]


def test_offline_pipeline_ok():
    rc = build_data.run(offline=True)
    assert rc == 0
    assert (L.DATA_DIR / "manifest.json").exists()
    assert (L.DATA_DIR / "update_log.json").exists()


def test_backup_and_restore(tmp_path):
    # backup_current crea last_valid; restore_last_valid lo recupera
    L.backup_current()
    assert (L.BACKUP_DIR / "last_valid.json").exists()
    assert L.restore_last_valid() is True


def test_apply_overrides_nulls_duplicate(tmp_path):
    from extract_legacy import load_macro, normalize
    macro = load_macro(ROOT / "legacy" / "dashboard-original.html")
    fresh = normalize(macro)
    # Crear un override de prueba y aplicarlo
    overrides_file = tmp_path / "overrides.json"
    overrides_file.write_text(json.dumps({
        "overrides": [{
            "indicator": "IMAI",
            "period": "Feb 26 P",
            "column": 0,
            "status": "revision",
            "value": None,
            "reason": "Test override",
        }]
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    log = L.apply_overrides(fresh, overrides_file)
    assert log, "debe registrar cambios de override"
    feb_imai = [o for o in fresh["indicators"]["IMAI"]["observations"] if o["period"].startswith("Feb 26")][0]
    assert feb_imai["values"][0] is None


def test_excel_has_new_sheets(tmp_path):
    build_calendar.main()  # calendario oficial disponible para la hoja de control
    build_excel.main()
    out = ROOT / "downloads" / "Indicadores_Macroeconomicos_Mexico_Actualizado.xlsx"
    assert out.exists()
    wb = openpyxl.load_workbook(out, read_only=True)
    for req in ("Síntesis de coyuntura", "Metodología y fuentes", "Control de actualizaciones"):
        assert req in wb.sheetnames
    assert "Resumen ejecutivo" not in wb.sheetnames
    # La hoja "Exportaciones" quedó obsoleta; BCMM se presenta como "Balanza comercial".
    assert "Exportaciones" not in wb.sheetnames
    # 23 hojas: 14 principales + divisiones (PIB en 2, DESOCUP en 2, EMIM en 2)
    # + 3 complementarios del Entorno financiero + 3 hojas fijas.
    assert len(wb.sheetnames) == 23
    # Las hojas principales y financieras clave deben existir.
    for orig in (
        "PIB oportuno", "Nivel PIB", "IGAE", "Balanza comercial", "INPC", "INPP",
        "Tasas laborales", "Población ocupada", "Formación bruta capital fijo",
        "IOAE", "EMIM (Manufactura)", "Subsectores EMIM", "EMOE (Confianza empresarial)",
        "IED", "Tipo de cambio FIX", "Tasa objetivo", "Reservas internacionales",
    ):
        assert orig in wb.sheetnames


def test_calendar_build_and_schema():
    rc = build_calendar.main()
    assert rc == 0
    cal = json.loads((L.DATA_DIR / "calendario_publicaciones.json").read_text(encoding="utf-8"))
    assert cal["items"], "el calendario debe tener publicaciones"
    required = {"clave", "indicador", "producto", "fecha_publicacion", "fecha_iso",
                "periodo_referencia", "frecuencia", "institucion", "estatus"}
    valid_statuses = {"publicado", "próximo", "pendiente", "no_anunciada", "evento", "regla"}
    for it in cal["items"]:
        assert required <= set(it), f"faltan campos en {it}"
        assert it["estatus"] in valid_statuses
    # fechas exactas ordenadas; items sin fecha (reglas) se dejan al final
    isos = [it["fecha_iso"] or "" for it in cal["items"]]
    assert isos == sorted(isos, key=lambda x: (x or "", 0 if x else 1))


def test_control_sheet_reflects_calendar():
    build_calendar.main()
    build_excel.main()
    out = ROOT / "downloads" / "Indicadores_Macroeconomicos_Mexico_Actualizado.xlsx"
    wb = openpyxl.load_workbook(out, read_only=True)
    ws = wb["Control de actualizaciones"]
    headers = [c.value for c in list(ws.iter_rows(min_row=4, max_row=4))[0]]
    assert "Próxima publicación (calendario)" in headers
