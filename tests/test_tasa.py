import json
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "indicadores.json"


def _load():
    return json.loads(DATA.read_text(encoding="utf-8"))["indicators"]["TASA"]


def test_tasa_nombre_y_no_tiie():
    ind = _load()
    assert "Tasa de Interes Interbancaria" in ind["nombre"]
    assert "objetivo" in ind["nombre"].lower()
    assert "TIIE" not in ind["nombre"]
    assert "no confundir con la tiie" in ind["descripcion"].lower()


def test_tasa_vigente_6_50():
    ind = _load()
    assert ind.get("tasa_vigente") == pytest.approx(6.5, rel=1e-6)
    assert ind["metrics"]["kpi"]["ultimoFmt"] == "6.50%"


def test_tasa_ultima_decision_sin_cambio():
    ind = _load()
    last = ind["policy_decisions"][-1]
    assert last["announcement_date"] == "2026-08-06"
    assert last["decision"] == "sin cambio"
    assert last["new_rate"] == pytest.approx(6.5, rel=1e-6)


def test_tasa_ultimo_ajuste_7_may_2026():
    ind = _load()
    adj = ind["ultimo_ajuste"]
    assert adj["announcement_date"] == "2026-05-07"
    assert adj["decision"] == "recorte"
    assert adj["change_bp"] == pytest.approx(-25.0, rel=1e-6)
    assert adj["effective_date"] == "2026-05-08"


def test_tasa_vigencia_8_may_2026():
    ind = _load()
    assert ind["vigente_desde"] == "2026-05-08"


def test_tasa_proxima_decision_24_sep_2026():
    ind = _load()
    cal = ind.get("calendario_publicaciones", [])
    prox = [e for e in cal if (e.get("status") or "").lower() in ("próximo", "proximo", "programado")]
    assert prox
    assert any((e.get("publication_date") or e.get("fecha_iso")) == "2026-09-24" for e in prox)


def test_tasa_no_cambio_diario_kpi():
    ind = _load()
    kpi = ind["metrics"]["kpi"]
    assert "cambio diario" not in kpi.get("varLabel", "").lower()
    assert kpi.get("varLabel") == "Último ajuste"


def test_tasa_no_maximo_minimo_kpi():
    ind = _load()
    kpi = ind["metrics"]["kpi"]
    assert "máximo" not in kpi or not kpi.get("maxFmt")
    assert "mínimo" not in kpi or not kpi.get("minFmt")


def test_tasa_resumen_bullets():
    ind = _load()
    resumen = ind["metrics"]["resumen"]
    assert len(resumen) <= 4
    assert any("6.50%" in b for b in resumen)
    assert any("6 de agosto de 2026" in b for b in resumen)
    assert any("7 de mayo de 2026" in b for b in resumen)
    assert any("24 de septiembre de 2026" in b for b in resumen)


def test_tasa_decisions_tiene_comunicados():
    ind = _load()
    decisions = ind["policy_decisions"]
    assert decisions
    for d in decisions:
        if d.get("comunicado_url"):
            assert d["comunicado_url"].startswith("https://")


def test_tasa_excel_tiene_dos_hojas():
    xlsx = ROOT / "downloads" / "indicadores" / "TASA" / "TASA_datos.xlsx"
    assert xlsx.exists(), f"No se generó {xlsx}"
    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    try:
        assert "Serie diaria" in wb.sheetnames
        assert "Decisiones" in wb.sheetnames
    finally:
        wb.close()
