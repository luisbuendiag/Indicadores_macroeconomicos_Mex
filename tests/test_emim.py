"""Tests para el rediseño de EMIM (18 columnas + subsectores)."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

import build_data
import build_excel
import lib_data as L
import lib_metrics as M
from sources import inegi_bulletin


PDF = Path("/tmp/emim2026_08.pdf")


def _emim_cols() -> list[dict]:
    return [
        {"label": "Producción (índice)", "index": 0, "fmt": "idx"},
        {"label": "Producción var. mensual original (%)", "index": 1, "fmt": "pct-frac"},
        {"label": "Producción var. anual original (%)", "index": 2, "fmt": "pct-frac"},
        {"label": "Producción var. mensual desest. (%)", "index": 3, "fmt": "pct-frac"},
        {"label": "Producción var. anual desest. (%)", "index": 4, "fmt": "pct-frac"},
        {"label": "Personal ocupado (índice)", "index": 5, "fmt": "idx"},
        {"label": "Personal var. mensual original (%)", "index": 6, "fmt": "pct-frac"},
        {"label": "Personal var. anual original (%)", "index": 7, "fmt": "pct-frac"},
        {"label": "Personal var. mensual desest. (%)", "index": 8, "fmt": "pct-frac"},
        {"label": "Personal var. anual desest. (%)", "index": 9, "fmt": "pct-frac"},
        {"label": "Horas trabajadas (índice)", "index": 10, "fmt": "idx"},
        {"label": "Horas var. mensual original (%)", "index": 11, "fmt": "pct-frac"},
        {"label": "Horas var. anual original (%)", "index": 12, "fmt": "pct-frac"},
        {"label": "Horas var. mensual desest. (%)", "index": 13, "fmt": "pct-frac"},
        {"label": "Horas var. anual desest. (%)", "index": 14, "fmt": "pct-frac"},
        {"label": "Remuneraciones medias reales (índice)", "index": 15, "fmt": "idx"},
        {"label": "Remuneraciones var. mensual desest. (%)", "index": 16, "fmt": "pct-frac"},
        {"label": "Remuneraciones var. anual desest. (%)", "index": 17, "fmt": "pct-frac"},
    ]


def _emim_ind() -> dict:
    return {
        "key": "EMIM",
        "nombre": "Encuesta Mensual de la Industria Manufacturera",
        "frecuencia": "Mensual",
        "unidad": "Índice base 2018=100",
        "columns": _emim_cols(),
        "observations": [
            {
                "period": "May 26",
                "values": [
                    100.0, 0.01, 0.02, 0.005, 0.02,
                    95.0, -0.001, 0.01, 0.0, 0.01,
                    98.0, 0.0, 0.015, 0.0, 0.015,
                    105.0, 0.0, 0.02,
                ],
            },
            {
                "period": "Jun 26",
                "values": [
                    101.0, 0.008, 0.018, 0.003, 0.018,
                    94.0, -0.002, 0.008, -0.001, 0.008,
                    98.5, 0.0, 0.012, 0.0, 0.012,
                    105.5, 0.0, 0.018,
                ],
            },
        ],
    }


def _emim_subsectores_detalle() -> dict:
    return {
        "311": {
            "nombre": "Industria alimentaria",
            "produccion_index": 102.0,
            "produccion_anual": 0.02,
            "personal_index": 100.0,
            "personal_anual": 0.0,
            "horas_index": 99.0,
            "horas_anual": -0.01,
            "remuneraciones_index": 110.0,
            "remuneraciones_anual": 0.03,
        },
        "312": {
            "nombre": "Industria de las bebidas y del tabaco",
            "produccion_index": 107.0,
            "produccion_anual": 0.009,
            "personal_index": 116.0,
            "personal_anual": 0.001,
            "horas_index": 122.0,
            "horas_anual": 0.018,
            "remuneraciones_index": 105.0,
            "remuneraciones_anual": 0.001,
        },
    }


@pytest.mark.skipif(not PDF.exists(), reason="PDF de prueba no disponible")
@pytest.mark.skipif(inegi_bulletin.pdfplumber is None, reason="pdfplumber no instalado")
def test_emim_parser_extracts_18_columns_and_subsectores():
    pdf = PDF.read_bytes()
    parsed = inegi_bulletin._parse_emim(pdf, (2026, 8, 1))
    assert parsed is not None
    assert "subsectores" in parsed
    for sub in (
        "produccion_index",
        "produccion_mensual_desest",
        "produccion_anual_desest",
        "produccion_anual_orig",
        "personal_index",
        "personal_mensual_desest",
        "personal_anual_desest",
        "personal_anual_orig",
        "horas_index",
        "horas_mensual_desest",
        "horas_anual_desest",
        "horas_anual_orig",
        "remuneraciones_index",
        "remuneraciones_mensual_desest",
        "remuneraciones_anual_desest",
    ):
        assert sub in parsed
        assert parsed[sub]
        assert parsed[sub][0]["value"] is not None

    for code, info in parsed["subsectores"].items():
        assert isinstance(code, str) and code
        assert isinstance(info, dict)
        for k in (
            "nombre",
            "produccion_index",
            "produccion_anual",
            "personal_index",
            "personal_anual",
            "horas_index",
            "horas_anual",
            "remuneraciones_index",
            "remuneraciones_anual",
        ):
            assert k in info


def test_compute_emim_metrics_pads_and_normalizes_subsectores():
    ind = _emim_ind()
    # Truncar a 5 columnas para forzar el padding.
    ind["columns"] = ind["columns"][:5]
    for o in ind["observations"]:
        o["values"] = o["values"][:5]
    ind["subsectores"] = _emim_subsectores_detalle()
    payload = {"indicators": {"EMIM": ind}}
    changes = build_data.compute_emim_metrics(payload)
    assert any("18 columnas" in c for c in changes)
    assert len(ind["columns"]) == 18
    assert all(len(o["values"]) == 18 for o in ind["observations"])
    # Subsectores se normaliza a plano.
    assert isinstance(ind["subsectores"], dict)
    assert any("Industria alimentaria" in k for k in ind["subsectores"])
    assert "subsectores_detalle" in ind


def test_emim_metrics_has_cards_and_resumen():
    ind = _emim_ind()
    ind["subsectores"] = _emim_subsectores_detalle()
    build_data.compute_emim_metrics({"indicators": {"EMIM": ind}})
    kpicfg = {
        "EMIM": {
            "art": "la",
            "noun": "Encuesta Mensual de la Industria Manufacturera",
            "grupo": "growth",
            "goodSign": 1,
            "vg": "f",
            "comp": "frente al mes previo",
            "ctx": " (índice base 2018=100)",
        }
    }
    metrics = M._emim_metrics(ind, kpicfg)
    assert metrics is not None
    assert "kpi" in metrics
    assert "annualVar" in metrics
    assert "resumen" in metrics
    assert "cards" in metrics["kpi"]
    assert len(metrics["kpi"]["cards"]) == 4
    for i, name in enumerate(("Producción", "Personal ocupado", "Horas trabajadas", "Remuneraciones medias reales")):
        c = metrics["kpi"]["cards"][i]
        assert c["name"] == name
        assert c["idxText"] != "—"
        # Remuneraciones no tiene variaciones originales (sólo desestacionalizadas).
        if name != "Remuneraciones medias reales":
            assert c["origYoyText"] != "—"
    assert metrics["kpi"]["ultimoP"] == "Jun 26"
    assert any("producción" in b.lower() for b in metrics["resumen"])


def test_build_emim_workbook_writes_two_sheets(tmp_path):
    ind = _emim_ind()
    ind["subsectores"] = _emim_subsectores_detalle()
    out = tmp_path / "EMIM_datos.xlsx"
    build_excel.build_emim_workbook(ind, out)
    assert out.exists()
    wb = openpyxl.load_workbook(out, read_only=True)
    assert "Indicadores" in wb.sheetnames
    assert "Subsectores" in wb.sheetnames


def test_validate_emim_18_columns():
    import validate as V
    ind = _emim_ind()
    payload = {"indicators": {"EMIM": ind}}
    errors, warnings = V.validate(payload)
    assert not errors
    assert not any("18 columnas" in w for w in warnings)
