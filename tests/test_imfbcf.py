import json
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "indicadores.json"


def _imfbcf():
    payload = json.loads(DATA.read_text(encoding="utf-8"))
    return payload["indicators"]["IMFBCF"]


def test_imfbcf_schema_40_columnas():
    ind = _imfbcf()
    assert len(ind["columns"]) == 40, f"IMFBCF debe tener 40 columnas, tiene {len(ind['columns'])}"
    for i, col in enumerate(ind["columns"]):
        assert col["index"] == i


def test_imfbcf_columnas_ordenadas():
    ind = _imfbcf()
    labels = [c["label"] for c in ind["columns"]]
    assert "índice" in labels[0].lower()
    assert "var. mensual" in labels[1].lower()
    assert "var. anual" in labels[2].lower()
    # Total original: col 25-27
    assert "índice" in labels[25].lower()
    assert "var. anual original" in labels[26].lower()
    assert "acumulado" in labels[27].lower()


def test_imfbcf_datos_recientes_completos():
    ind = _imfbcf()
    obs = ind["observations"][-1]
    assert len(obs["values"]) == 40
    assert obs["values"][0] is not None
    assert obs["values"][1] is not None
    assert obs["values"][2] is not None


def test_imfbcf_kpi_resumen_sin_favorable_adverso():
    ind = _imfbcf()
    resumen = ind.get("metrics", {}).get("resumen", [])
    assert resumen, "IMFBCF debe tener resumen"
    texto = " ".join(resumen).lower()
    assert "favorable" not in texto
    assert "adverso" not in texto
    assert "se clasifica" not in texto


def test_imfbcf_kpi_tiene_acum():
    ind = _imfbcf()
    kpi = ind.get("metrics", {}).get("kpi", {})
    assert kpi.get("acumText") is not None
    assert kpi.get("acumText") != "—"


def test_imfbcf_cifra_actual():
    ind = _imfbcf()
    kpi = ind.get("metrics", {}).get("kpi", {})
    ult = kpi.get("ultimoRaw")
    assert ult is not None
    assert 50 < ult < 130


def test_imfbcf_cards_componentes():
    ind = _imfbcf()
    kpi = ind.get("metrics", {}).get("kpi", {})
    cards = kpi.get("cards", [])
    assert len(cards) >= 8, f"Esperaba al menos 8 cards, hay {len(cards)}"
    for c in cards:
        assert "nivelText" in c
        assert "yoyText" in c


def test_imfbcf_url_boletin_oficial():
    ind = _imfbcf()
    url = ind.get("url_boletin_oficial")
    assert url
    assert "inegi.org.mx/contenidos/saladeprensa/boletines" in url
    assert url.endswith(".pdf")


def test_imfbcf_excel_individual_hojas(tmp_path):
    import sys
    sys.path.insert(0, str(ROOT / "scripts"))
    from build_excel import _build_imfbcf_workbook
    ind = _imfbcf()
    out = tmp_path / "IMFBCF_datos.xlsx"
    _build_imfbcf_workbook(ind, out)
    assert out.exists()

    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert "Niveles" in wb.sheetnames
    assert "Variaciones" in wb.sheetnames
    assert "Originales" in wb.sheetnames
    assert "Resumen" in wb.sheetnames


def test_imfbcf_fuente_banxico_inegi():
    ind = _imfbcf()
    fuente = ind.get("fuente", {})
    assert "Banco de México" in fuente.get("nombre", "") or "INEGI" in fuente.get("nombre", "")
    assert fuente.get("link")


def test_imfbcf_variaciones_componentes_ultimo_periodo():
    ind = _imfbcf()
    obs = ind["observations"][-1]
    # Construcción y maquinaria deben tener mensual y anual
    assert obs["values"][4] is not None
    assert obs["values"][5] is not None
    assert obs["values"][7] is not None
    assert obs["values"][8] is not None
