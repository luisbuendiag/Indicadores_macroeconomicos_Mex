"""Regresiones para frescura, métricas, productos individuales y metadatos."""
from __future__ import annotations

import json
import zipfile
from datetime import datetime
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "indicadores.json"


@pytest.fixture
def payload():
    return json.loads(DATA.read_text(encoding="utf-8"))


def test_meta_last_update_ct_mexico_city(payload):
    meta = payload["meta"]
    assert "last_update_ct" in meta
    # Debe ser un datetime con zona horaria explícita.
    dt = datetime.fromisoformat(meta["last_update_ct"])
    assert dt.tzinfo is not None
    # El offset de America/Mexico_City es -06:00 en verano o -05:00 en invierno.
    offset = dt.utcoffset()
    assert offset is not None
    assert -8 * 3600 < offset.total_seconds() <= -5 * 3600


def test_periodo_referencia_reciente_from_principal(payload):
    ref = payload["meta"].get("periodo_referencia_reciente")
    assert ref and ref.get("period")
    # Debe pertenecer a un indicador principal.
    from lib_kpicfg import get_cfg
    principal = get_cfg("PRINCIPAL")
    matches = [ind for k, ind in payload["indicators"].items() if k in principal and ind.get("last_observation") == ref["period"]]
    assert matches, f"El periodo de referencia reciente {ref['period']} no pertenece a un indicador principal"


def test_freshness_states_defined():
    from lib_freshness import ESTADOS
    values = set(ESTADOS.values())
    for st in ("ACTUALIZADO", "PUBLICACIÓN PENDIENTE", "REZAGADO", "ERROR DE FUENTE"):
        assert st in values


def test_igae_estado_actualizado(payload):
    ind = payload["indicators"]["IGAE"]
    assert ind["estado"] == "ACTUALIZADO"
    assert ind["periodo_referencia_oficial"]
    assert ind["fecha_publicacion_oficial"]


def test_inpc_estado_actualizado(payload):
    ind = payload["indicators"]["INPC"]
    assert ind["estado"] == "ACTUALIZADO"


def test_consumo_estado_rezagado(payload):
    ind = payload["indicators"]["CONSUMO"]
    assert ind["estado"] == "REZAGADO"
    assert ind.get("motivo_frescura")


def test_metrics_present_and_resumen_length(payload):
    for key in ("IGAE", "INPC"):
        metrics = payload["indicators"][key].get("metrics")
        assert metrics, f"{key} no tiene métricas"
        assert metrics.get("kpi") is not None
        resumen = metrics.get("resumen", [])
        assert 3 <= len(resumen) <= 4, f"{key} resumen debe tener 3-4 bullets"


def test_url_boletin_oficial_set(payload):
    for key in ("IGAE", "INPC"):
        ind = payload["indicators"][key]
        assert ind.get("url_boletin_oficial"), f"{key} sin URL de boletín"


def test_individual_excel_exists_and_has_columns(payload):
    import openpyxl
    for key in ("IGAE", "INPC"):
        ind = payload["indicators"][key]
        assert ind.get("xlsx_disponible"), f"{key} no marcó xlsx_disponible"
        path = ROOT / ind["url_excel_individual"]
        assert path.exists(), f"{path} no existe"
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [ws.cell(5, c).value for c in range(1, ws.max_column + 1)]
        required = ["Periodo", "Fecha", "Nivel", "Variación mensual", "Variación trimestral", "Variación anual"]
        for col in required:
            assert col in headers, f"{key} falta columna {col}"


def test_individual_note_generated_for_pilot(payload):
    for key in ("IGAE", "INPC"):
        ind = payload["indicators"][key]
        assert ind.get("nota_disponible"), f"{key} nota no disponible: {ind.get('nota_causa')}"
        path = ROOT / ind["url_nota_individual"]
        assert path.exists(), f"{path} no existe"


def test_historical_data_not_truncated(payload):
    # El historial debe conservarse desde 2018 o el primer dato disponible.
    ind = payload["indicators"]["IGAE"]
    obs = ind["observations"]
    assert len(obs) >= 10, "IGAE debe tener al menos 10 observaciones para el piloto"
    # No debe haber duplicados.
    periods = [o["period"] for o in obs]
    assert len(periods) == len(set(periods)), "Periodos duplicados en IGAE"


def test_pilot_product_flags(payload):
    for key in ("IGAE", "INPC"):
        ind = payload["indicators"][key]
        assert ind.get("xlsx_disponible")
        assert ind.get("nota_disponible")
        assert ind.get("url_boletin_oficial")
