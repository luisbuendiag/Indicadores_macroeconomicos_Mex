"""Tests para el indicador RESERVAS (reservas internacionales)."""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

import lib_format as F
import lib_kpicfg

ROOT = Path(__file__).resolve().parents[1]
DATA_FILE = ROOT / "data" / "indicadores.json"
META_FILE = ROOT / "config" / "indicadores_meta.json"
XLSX_FILE = ROOT / "downloads" / "Indicadores_Macroeconomicos_Mexico_Actualizado.xlsx"


def _payload():
    return json.loads(DATA_FILE.read_text(encoding="utf-8"))


def _meta():
    return json.loads(META_FILE.read_text(encoding="utf-8"))


def test_reservas_is_complementario():
    """RESERVAS pertenece a Entorno financiero, no al Panorama."""
    meta = _meta()
    assert "RESERVAS" in meta["complementario"]
    assert "RESERVAS" not in meta["principal"]
    cfg = lib_kpicfg.get_cfg()
    assert "RESERVAS" in cfg["COMPLEMENTARIOS"]
    assert "RESERVAS" not in cfg["PRINCIPAL"]


def test_reservas_meta_columns():
    """El perfil define 4 columnas con los formatos correctos.

    RESERVAS está en scaffolds (por convención del repo) y es copiado al
    profile en runtime por lib_data.apply_profile.
    """
    meta = _meta()
    prof = meta["scaffolds"].get("RESERVAS") or meta["profile"].get("RESERVAS")
    assert prof is not None
    cols = prof["columns"]
    assert len(cols) == 4
    assert cols[0]["label"] == "Reserva internacional"
    assert cols[0]["fmt"] == "mdd"
    assert cols[1]["label"] == "Cambio semanal"
    assert cols[1]["fmt"] == "mdd-signed"
    assert cols[2]["label"] == "Cambio YTD"
    assert cols[2]["fmt"] == "mdd-signed"
    assert cols[3]["label"] == "Variación anual"
    assert cols[3]["fmt"] == "pct-frac"


def test_reservas_kpi_format_and_values():
    """El KPI de RESERVAS sigue el rediseño: mdd sin $, signos correctos, YTD dual."""
    payload = _payload()
    r = payload["indicators"]["RESERVAS"]
    k = r["metrics"]["kpi"]

    assert k["ultimoFmt"].endswith(" mdd")
    assert "$" not in k["ultimoFmt"]
    assert k["varText"].endswith(" mdd")
    assert k["varLabel"] == "Cambio semanal"
    assert "%" in k["yoyText"]
    assert k["yoyLabel"] == "Variación anual"
    assert k["acumText"].endswith(" mdd")
    assert k["acumLabel"] == "Cambio YTD"
    assert k["acumPctText"] is None or "%" in k["acumPctText"]
    assert k["ultimoP"].startswith("SEMANA AL")


def test_reservas_resumen_bullets():
    """El resumen genera 4 bullets coherentes."""
    payload = _payload()
    r = payload["indicators"]["RESERVAS"]
    resumen = r["metrics"]["resumen"]
    assert len(resumen) == 4
    assert "258,564 mdd" in resumen[0]
    assert "variación semanal" in resumen[1].lower()
    assert "tasa anual" in resumen[2].lower() or "anual" in resumen[2].lower()
    assert "acumulado" in resumen[3].lower()


def test_reservas_observations_have_four_columns():
    """Cada observación enriquecida tiene 4 valores y los derivados son correctos."""
    payload = _payload()
    r = payload["indicators"]["RESERVAS"]
    obs = r["observations"]
    assert all(len(o["values"]) == 4 for o in obs)
    # Primera observación: sin cambio semanal, sin YTD, sin variación anual.
    assert obs[0]["values"][1] is None
    assert obs[0]["values"][2] is None
    assert obs[0]["values"][3] is None
    # Última observación: todos los derivados presentes.
    last = obs[-1]
    assert last["values"][0] is not None
    assert last["values"][1] is not None
    assert last["values"][2] is not None
    assert last["values"][3] is not None
    # El cambio semanal es la diferencia con la observación previa.
    assert round(last["values"][0] - obs[-2]["values"][0], 6) == pytest.approx(last["values"][1])


def test_reservas_ytd_uses_previous_year_end():
    """Cambio YTD = saldo - última observación del año previo."""
    payload = _payload()
    obs = payload["indicators"]["RESERVAS"]["observations"]
    # Encontrar la primera observación de 2019.
    first_2019 = next(o for o in obs if o["period"].startswith("2019"))
    last_2018 = [o for o in obs if o["period"].startswith("2018-")][-1]
    if first_2019["values"][2] is not None:
        assert round(first_2019["values"][0] - last_2018["values"][0], 6) == pytest.approx(first_2019["values"][2])


def test_reservas_yoy_uses_52_week_lag():
    """Variación anual apunta a ~52 semanas atrás."""
    from datetime import date, timedelta
    payload = _payload()
    obs = payload["indicators"]["RESERVAS"]["observations"]
    last = obs[-1]
    last_date = date.fromisoformat(last["period"])
    target = last_date - timedelta(weeks=52)
    candidates = [o for o in obs if o["period"] <= target.isoformat()]
    assert candidates
    base = max(candidates, key=lambda o: o["period"])
    expected_pct = (last["values"][0] - base["values"][0]) / abs(base["values"][0])
    assert round(expected_pct, 6) == pytest.approx(last["values"][3])


def test_reservas_boletin_button():
    """El botón de producto apunta al estado de cuenta de Banxico."""
    payload = _payload()
    r = payload["indicators"]["RESERVAS"]
    assert r.get("boletin_label") == "ESTADO DE CUENTA"
    assert "banxico.org.mx" in r.get("url_boletin_oficial", "")
    assert "reservas-internacionales" in r.get("url_boletin_oficial", "")


def test_reservas_format_lib():
    """fmt_val soporta mdd y mdd-signed con signos Unicode."""
    assert F.fmt_val(258564.1, "mdd") == "258,564 mdd"
    assert F.fmt_val(1192, "mdd-signed") == "+1,192 mdd"
    assert F.fmt_val(-45, "mdd-signed") == "−45 mdd"
    assert F.fmt_val(0.057959, "pct-frac") == "5.8%"
    assert F.fmt_val(-0.012, "pct-frac") == "-1.2%"
    # El KPI agrega el signo + en los textos del resumen.
    assert F.fmt_val(0.057959, "pct-frac").startswith("5.8")
    assert F.fmt_val(-0.012, "pct-frac").startswith("-1.2")


def test_reservas_excel_sheet():
    """El libro Excel contiene la hoja de reservas con 4 columnas de datos."""
    wb = openpyxl.load_workbook(XLSX_FILE)
    assert "Reservas internacionales" in wb.sheetnames
    ws = wb["Reservas internacionales"]
    headers = [ws.cell(4, c).value for c in range(1, ws.max_column + 1)]
    assert headers == ["Periodo", "Reserva internacional", "Cambio semanal", "Cambio YTD", "Variación anual"]
