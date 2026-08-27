import json
import re
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "indicadores.json"
META = ROOT / "config" / "indicadores_meta.json"
APP = (ROOT / "assets" / "js" / "app.js").read_text(encoding="utf-8")
CONFIG = (ROOT / "assets" / "js" / "config.js").read_text(encoding="utf-8")


def _imcp():
    payload = json.loads(DATA.read_text(encoding="utf-8"))
    return payload["indicators"]["CONSUMO"]


def _meta():
    return json.loads(META.read_text(encoding="utf-8"))


def test_imcp_es_principal():
    meta = _meta()
    assert "CONSUMO" in meta["principal"]
    ind = _imcp()
    assert ind["clasificacion"] == "principal"


def test_imcp_nombre_oficial_y_sigla():
    ind = _imcp()
    assert ind["nombre"] == "Indicador Mensual del Consumo Privado"
    assert ind.get("sigla") == "IMCP"


def test_imcp_descripcion_incluye_base_y_frecuencia():
    ind = _imcp()
    desc = ind["descripcion"].lower()
    assert "base 2018" in desc
    assert "frecuencia mensual" in desc


def test_imcp_37_columnas_ordenadas():
    ind = _imcp()
    assert len(ind["columns"]) == 37, f"IMCP debe tener 37 columnas, tiene {len(ind['columns'])}"
    for i, col in enumerate(ind["columns"]):
        assert col["index"] == i
    labels = [c["label"] for c in ind["columns"]]
    assert "índice" in labels[0].lower()
    assert labels[1].lower().startswith("var. mensual")
    assert labels[2].lower().startswith("var. anual")
    assert "acumulado" in labels[4].lower()
    # Componentes principales
    for concept in ("nacional", "bienes", "servicios", "importado", "duradero"):
        assert any(concept in l.lower() for l in labels), f"falta componente {concept}"


def test_imcp_no_doble_multiplicacion_por_100():
    """Las variaciones porcentuales se almacenan como fracciones (<2.0 = 200%)."""
    ind = _imcp()
    pct_cols = [i for i, c in enumerate(ind["columns"]) if c["fmt"] == "pct-frac"]
    for o in ind["observations"]:
        for i in pct_cols:
            v = o["values"][i]
            if v is not None:
                assert abs(v) < 2.0, f"valor sospechosamente grande en {o['period']} col{i}: {v}"


def test_imcp_kpi_textos_con_un_solo_porcentaje():
    ind = _imcp()
    kpi = ind["metrics"]["kpi"]
    for txt in (kpi["varText"], kpi["yoyText"]):
        assert "%" in txt
        assert "%%" not in txt
    assert kpi["varText"].startswith("+") or kpi["varText"].startswith("-") or kpi["varText"] == "—"


def test_imcp_ultimo_periodo_todos_los_componentes():
    ind = _imcp()
    obs = ind["observations"][-1]
    assert len(obs["values"]) == 37
    # Agregado y componentes principales
    for i in (0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14):
        assert obs["values"][i] is not None, f"falta valor col {i} en {obs['period']}"
    # Durabilidad
    for i in range(25, 37):
        assert obs["values"][i] is not None, f"falta durabilidad col {i} en {obs['period']}"


def test_imcp_index_rango_realista():
    ind = _imcp()
    kpi = ind["metrics"]["kpi"]
    assert 50 < kpi["ultimoRaw"] < 130


def test_imcp_variaciones_no_nulas_ultimo():
    ind = _imcp()
    obs = ind["observations"][-1]
    assert obs["values"][0] is not None  # índice
    assert obs["values"][1] is not None  # mensual
    assert obs["values"][2] is not None  # anual
    assert obs["values"][4] is not None  # acumulado


def test_imcp_resumen_max_4_bullets():
    ind = _imcp()
    resumen = ind["metrics"]["resumen"]
    assert resumen
    assert len(resumen) <= 4


def test_imcp_mantiene_historial_disponible():
    ind = _imcp()
    # La serie debe conservar el historial disponible en los boletines oficiales,
    # sin eliminar años enteros artificialmente.
    assert len(ind["observations"]) >= 12
    first_year = ind["observations"][0]["period"].split()[-1]
    assert int(first_year) <= 25, f"el inicio de la serie es muy reciente: {first_year}"


def test_imcp_filtros_ventanas():
    ind = _imcp()
    wins = ind["windows"]
    ids = [w["id"] for w in wins]
    assert "3m" in ids and "6m" in ids and "1a" in ids and "5a" in ids and "max" in ids


def test_imcp_url_boletin_oficial():
    ind = _imcp()
    url = ind.get("url_boletin_oficial")
    assert url
    assert "inegi.org.mx/contenidos/saladeprensa/boletines" in url
    assert "imcp" in url
    assert url.endswith(".pdf")


def test_imcp_fuente_usa_boletin_pdf_no_ids_bie():
    ind = _imcp()
    fuente = ind.get("fuente", {})
    assert "pdf" in fuente.get("metodo", "").lower() or "boletín" in fuente.get("metodo", "").lower()
    # La serie es la plantilla del boletín, no un ID numérico de BIE hardcodeado.
    assert "imcpmi" in fuente.get("serie", "")


def test_imcp_excel_hoja_agregada():
    from openpyxl import load_workbook
    wb = load_workbook(ROOT / "downloads" / "Indicadores_Macroeconomicos_Mexico_Actualizado.xlsx", read_only=True)
    assert "Consumo Privado" in wb.sheetnames


def test_imcp_excel_individual_existe():
    path = ROOT / "downloads" / "indicadores" / "CONSUMO" / "CONSUMO_datos.xlsx"
    assert path.exists()


def test_imcp_ui_tiene_desglose_en_app():
    assert 'ind.key === "CONSUMO"' in APP
    assert "Desempeño por origen y durabilidad" in APP
    assert "Bienes duraderos nacionales" in APP
    assert "Bienes importados" in APP


def test_imcp_kpicfg_tiene_acumulado():
    m = re.search(r'CONSUMO:\s*\{[^}]*\}', CONFIG)
    assert m, "no se encontró KPICFG de CONSUMO"
    cfg = m.group(0)
    assert 'acumCol: 4' in cfg
    assert 'acumFmt: "pct-frac"' in cfg
    assert 'acumLabel: "Acumulado ene-mes"' in cfg
