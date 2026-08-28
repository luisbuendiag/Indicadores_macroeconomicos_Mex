import json
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data" / "indicadores.json"


def _ioae():
    payload = json.loads(DATA.read_text(encoding="utf-8"))
    return payload["indicators"]["IOAE"]


def test_ioae_schema_13_columnas():
    ind = _ioae()
    assert len(ind["columns"]) == 13, f"IOAE debe tener 13 columnas, tiene {len(ind['columns'])}"
    for i, col in enumerate(ind["columns"]):
        assert col["index"] == i


def test_ioae_columnas_ordenadas():
    ind = _ioae()
    labels = [c["label"].lower() for c in ind["columns"]]
    assert "igae" in labels[0] and "anual" in labels[0]
    assert "ic95" in labels[1] and "inferior" in labels[1]
    assert "ic95" in labels[2] and "superior" in labels[2]
    assert "igae" in labels[3] and "mensual" in labels[3]
    assert "secundarias" in labels[4]
    assert "terciarias" in labels[7]
    assert "fecha" in labels[10]
    assert "carácter" in labels[11]
    assert "igae observado" in labels[12]


def test_ioae_datos_recientes_completos():
    ind = _ioae()
    obs = ind["observations"][-1]
    assert len(obs["values"]) == 13
    assert obs["values"][0] is not None
    assert obs["values"][1] is not None
    assert obs["values"][2] is not None
    assert obs["values"][3] is not None


def test_ioae_kpi_tiene_ic_y_error():
    ind = _ioae()
    kpi = ind.get("metrics", {}).get("kpi", {})
    assert kpi.get("ultimoRaw") is not None
    assert kpi.get("icWidthText") is not None
    # El error y el RMSE solo existen cuando ya hay IGAE observado disponible
    if kpi.get("observedRaw") is not None:
        assert kpi.get("errorText") is not None
        assert kpi.get("rmseText") is not None


def test_ioae_cifra_actual_fraccion():
    ind = _ioae()
    kpi = ind.get("metrics", {}).get("kpi", {})
    ult = kpi.get("ultimoRaw")
    assert ult is not None
    assert -0.5 < ult < 0.5  # fracciones +/- 50%


def test_ioae_url_boletin_oficial():
    ind = _ioae()
    url = ind.get("url_boletin_oficial")
    assert url
    assert "inegi.org.mx/contenidos/saladeprensa/boletines" in url
    assert url.endswith(".pdf")


def test_ioae_excel_individual_hoja(tmp_path):
    import sys
    sys.path.insert(0, str(ROOT / "scripts"))
    from build_excel import _build_individual_workbook, get_cfg
    ind = _ioae()
    out = tmp_path / "IOAE_datos.xlsx"
    _build_individual_workbook(ind, get_cfg(), get_cfg("KPICFG"), out)
    assert out.exists()

    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert "IOAE" in wb.sheetnames


def test_ioae_igae_observado_copia_cuando_existe():
    ind = _ioae()
    igae = json.loads(DATA.read_text(encoding="utf-8"))["indicators"].get("IGAE")
    if not igae or not igae.get("observations"):
        pytest.skip("IGAE sin observaciones")
    igae_ym = {inegi_label_to_ym(o["period"]): o["values"][2] for o in igae["observations"] if len(o["values"]) > 2}
    coinciden = 0
    for o in ind["observations"]:
        ym = inegi_label_to_ym(o["period"])
        if ym in igae_ym and o["values"][12] == igae_ym[ym]:
            coinciden += 1
    if not igae_ym:
        pytest.skip("IGAE no tiene variaciones anuales para contrastar")
    assert coinciden > 0, "Debe haber al menos un IOAE con IGAE observado copiado"


def inegi_label_to_ym(label: str) -> str:
    """Convierte 'Abr 26' -> '2026-04'."""
    meses = {
        "ene": "01", "feb": "02", "mar": "03", "abr": "04", "may": "05", "jun": "06",
        "jul": "07", "ago": "08", "sep": "09", "oct": "10", "nov": "11", "dic": "12",
    }
    partes = label.split()
    if len(partes) != 2:
        return ""
    m = meses.get(partes[0].lower(), "")
    if not m:
        return ""
    yy = int(partes[1])
    year = 2000 + yy
    return f"{year}-{m}"
