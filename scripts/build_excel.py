"""Genera downloads/Indicadores_Macroeconomicos_Mexico_Actualizado.xlsx.

Parte del Excel base (data/source/Indicadores_base.xlsx), conserva sus hojas
originales SIN alterarlas innecesariamente y agrega tres hojas nuevas:
  - "Síntesis de coyuntura"
  - "Metodología y fuentes"
  - "Control de actualizaciones"

No usa rutas locales ni fórmulas frágiles: los cálculos se resuelven con
fórmulas de Excel autocontenidas cuando aportan (promedios, máximos), sobre
rangos internos del propio libro.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import lib_data as L
from lib_kpicfg import get_cfg
from lib_metrics import annual_var, compute_var, fmt_val, primary_series
from sources import inegi

ROOT = Path(__file__).resolve().parents[1]
BASE_XLSX = ROOT / "data" / "source" / "Indicadores_base.xlsx"
OUT_DIR = ROOT / "downloads"
OUT_XLSX = OUT_DIR / "Indicadores_Macroeconomicos_Mexico_Actualizado.xlsx"

DKGREEN = "FF002F2A"
GREEN = "FF1E5B4F"
GOLD = "FFA57F2C"
LINE = "FFE6E0D2"
PAPER = "FFF7F5EF"

TITLE = Font(name="Calibri", size=15, bold=True, color=DKGREEN)
H = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
LBL = Font(name="Calibri", size=10, bold=True, color=DKGREEN)
TXT = Font(name="Calibri", size=10, color="FF161A1D")
MUT = Font(name="Calibri", size=9, color="FF6C6F6A", italic=True)
HEAD_FILL = PatternFill("solid", fgColor=DKGREEN)
BAND_FILL = PatternFill("solid", fgColor=PAPER)
THIN = Side(style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

IND_DIR = ROOT / "downloads" / "indicadores"
PILOT_KEYS = ["IGAE", "INPC"]


def _period_date(period: str) -> date | None:
    """Convierte una etiqueta de periodo (ej. 'Abr 26 P', '1T-26') en una fecha de inicio."""
    ym = inegi.label_to_ym(period)
    if not ym:
        return None
    return date.fromisoformat(ym + "-01")


def _caracter_dato(period: str) -> str:
    p = period.strip()
    if p.endswith(" P"):
        return "Preliminar"
    if p.endswith(" R"):
        return "Revisado"
    return "Definitivo"


def _xl_fmt(fmt: str) -> str:
    """Formato numérico de Excel según el descriptor de config."""
    # Variaciones fraccionarias (0.015 = 1.5%) se muestran con % nativo de Excel.
    # Valores ya en puntos porcentuales (4.5) usan % literal para no multiplicar.
    if fmt == "num" or fmt == "usd":
        return "#,##0"
    if fmt == "bill":
        return "#,##0.00"
    if fmt == "pct-frac":
        return "0.0%"
    if fmt == "pct-raw":
        return '0.0"%"'
    if fmt == "idx":
        return "#,##0.0"
    if fmt == "fx":
        return "$#,##0.00"
    return "#,##0.0"


def _prev_valid_i(idxs: list[int], i: int) -> int | None:
    """Devuelve el índice válido inmediatamente anterior a i."""
    for j in range(i - 1, -1, -1):
        if j in idxs:
            return j
    return None


def _bucket_variation(var_info: dict | None, yoy_info: dict | None, freq: str, cfg: dict) -> dict[str, str]:
    """Asigna var_text e yoy_text a las columnas mensual/trimestral/anual."""
    buckets = {"mensual": None, "trimestral": None, "anual": None}
    freq_l = (freq or "").lower()

    def place(info, label):
        if not info or not info.get("text") or info["text"] == "—":
            return
        lbl = (label or "").lower()
        if "anual" in lbl and "trimestral" not in lbl:
            buckets["anual"] = info["text"]
        elif "trimestral" in lbl:
            buckets["trimestral"] = info["text"]
        elif "mensual" in lbl or "mes" in lbl:
            buckets["mensual"] = info["text"]
        elif "semanal" in lbl or "semana" in lbl:
            # Agrupación mensual: conservamos la lectura del periodo en la columna mensual.
            if "mensual" in freq_l:
                buckets["mensual"] = info["text"]
        elif "diaria" in lbl or "diario" in lbl:
            if "mensual" in freq_l:
                buckets["mensual"] = info["text"]
        elif freq_l.startswith("trimestral"):
            buckets["trimestral"] = info["text"]
        elif "mensual" in freq_l:
            buckets["mensual"] = info["text"]

    # Variación principal
    if cfg.get("varCol") is not None:
        place(var_info, cfg.get("varLabel"))
    elif cfg.get("varMode") == "pct-yoy":
        if var_info and var_info.get("text"):
            buckets["anual"] = var_info["text"]
    elif "trimestral" in freq_l:
        buckets["trimestral"] = var_info["text"] if var_info and var_info.get("text") and var_info["text"] != "—" else None
    elif "mensual" in freq_l or "semanal" in freq_l or "diaria" in freq_l or "diario" in freq_l:
        buckets["mensual"] = var_info["text"] if var_info and var_info.get("text") and var_info["text"] != "—" else None

    # Variación secundaria (anual / trimestral)
    if yoy_info:
        place(yoy_info, cfg.get("yoyLabel"))
    return buckets


def _build_pibsec_workbook(ind: dict, out_path: Path):
    """Genera el Excel individual de PIBSEC con dos hojas claras: Niveles y Variaciones."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    title = ind.get("nombre", "PIBSEC")
    src = ind.get("fuente", {}).get("nombre", "INEGI")
    url = ind.get("url_boletin_oficial") or ind.get("fuente", {}).get("link") or "—"
    pub = ind.get("fecha_publicacion") or "—"
    header_note = f"Fuente: {src} · Frecuencia: Trimestral · Unidad: millones de pesos a precios de 2018 · Boletín: {url} · Fecha de publicación: {pub}"

    # Mapeos de columnas para PIBSEC: 0-2 niveles actividades, 3-4 terciarias var, 5 nivel PIB, 6-11 var total/prim/sec.
    def _make_sheet(sheet_name, include_cols, billions=False):
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws["A1"] = f"{title} — {sheet_name}"
        ws["A1"].font = TITLE
        ws["A2"] = header_note
        ws["A2"].font = MUT
        ws["A3"] = "Los niveles se presentan en billones de pesos de 2018 (valor / 1,000,000) y las variaciones en porcentaje."
        ws["A3"].font = MUT

        headers = ["Periodo", "Fecha"] + [ind["columns"][c]["label"] for c in include_cols]
        r0 = 5
        for i, h in enumerate(headers, start=1):
            ws.cell(row=r0, column=i, value=h)
        _style_header(ws, r0, len(headers))

        r = r0 + 1
        for o in ind.get("observations", []):
            d = _period_date(o["period"])
            ws.cell(row=r, column=1, value=o["period"]).font = TXT
            ws.cell(row=r, column=2, value=d).font = TXT
            if d:
                ws.cell(row=r, column=2).number_format = "yyyy-mm-dd"
            for j, col_i in enumerate(include_cols, start=3):
                v = o["values"][col_i] if col_i < len(o["values"]) else None
                cell = ws.cell(row=r, column=j, value=(v / 1e6 if v is not None and billions else v))
                cell.font = TXT
                if v is not None:
                    if billions:
                        cell.number_format = '#,##0.00" billones de pesos de 2018"'
                    else:
                        cell.number_format = _xl_fmt(ind["columns"][col_i].get("fmt", "num"))
                if r % 2 == 0:
                    cell.fill = BAND_FILL
            r += 1
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
        _autow(ws, [16, 13] + [26] * len(include_cols))

    _make_sheet("Niveles", [5, 0, 1, 2], billions=True)
    _make_sheet("Variaciones", [6, 7, 8, 9, 10, 11, 3, 4], billions=False)
    wb.save(out_path)


def _build_imai_workbook(ind: dict, out_path: Path) -> None:
    """Genera el Excel individual de IMAI con tres hojas: Niveles, Variaciones y Resumen."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    title = ind.get("nombre", "IMAI")
    src = ind.get("fuente", {}).get("nombre", "INEGI")
    url = ind.get("url_boletin_oficial") or ind.get("fuente", {}).get("link") or "—"
    pub = ind.get("fecha_publicacion") or "—"
    header_note = f"Fuente: {src} · Frecuencia: Mensual · Unidad: índice base 2018=100 / % · Boletín: {url} · Fecha de publicación: {pub}"

    def _make_sheet(sheet_name, include_cols, note):
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws["A1"] = f"{title} — {sheet_name}"
        ws["A1"].font = TITLE
        ws["A2"] = header_note
        ws["A2"].font = MUT
        ws["A3"] = note
        ws["A3"].font = MUT

        headers = ["Periodo", "Fecha"] + [ind["columns"][c]["label"] for c in include_cols]
        r0 = 5
        for i, h in enumerate(headers, start=1):
            ws.cell(row=r0, column=i, value=h)
        _style_header(ws, r0, len(headers))

        r = r0 + 1
        for o in ind.get("observations", []):
            d = _period_date(o["period"])
            ws.cell(row=r, column=1, value=o["period"]).font = TXT
            ws.cell(row=r, column=2, value=d).font = TXT
            if d:
                ws.cell(row=r, column=2).number_format = "yyyy-mm-dd"
            for j, col_i in enumerate(include_cols, start=3):
                v = o["values"][col_i] if col_i < len(o["values"]) else None
                cell = ws.cell(row=r, column=j, value=v)
                cell.font = TXT
                if v is not None:
                    cell.number_format = _xl_fmt(ind["columns"][col_i].get("fmt", "num"))
                if r % 2 == 0:
                    cell.fill = BAND_FILL
            r += 1
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
        _autow(ws, [16, 13] + [24] * len(include_cols))

    _make_sheet(
        "Niveles", [0, 6, 7, 8, 9],
        "Los niveles son índices desestacionalizados con base 2018=100."
    )
    _make_sheet(
        "Variaciones",
        [1, 2, 14, 15, 16, 17, 10, 11, 12, 13],
        "Las variaciones son porcentajes: mensuales (cols 1 y 14-17) y anuales (cols 2 y 10-13), cifras desestacionalizadas."
    )

    # Resumen con los KPIs y componentes del último periodo.
    ws = wb.create_sheet("Resumen")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{title} — Resumen"
    ws["A1"].font = TITLE
    ws["A2"] = header_note
    ws["A2"].font = MUT
    ws["A3"] = "Lectura del periodo más reciente. Los porcentajes son cifras desestacionalizadas; el acumulado es original."
    ws["A3"].font = MUT

    metrics = ind.get("metrics", {})
    kpi = metrics.get("kpi", {})
    r0 = 5
    headers = ["Concepto", "Valor", "Periodo"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=r0, column=i, value=h)
    _style_header(ws, r0, len(headers))

    rows = []
    if kpi:
        rows.append(["Índice", kpi.get("ultimoFmt"), kpi.get("ultimoP")])
        rows.append(["Var. mensual desest.", kpi.get("varText"), kpi.get("ultimoP")])
        rows.append(["Var. anual desest.", kpi.get("yoyText"), kpi.get("ultimoP")])
        rows.append(["Acumulado ene-mes", kpi.get("acumText"), kpi.get("ultimoP")])
    for c in kpi.get("cards", []):
        rows.append([c["name"] + " — índice", c.get("nivelText"), kpi.get("ultimoP")])
        rows.append([c["name"] + " — var. mensual", c.get("momText"), kpi.get("ultimoP")])
        rows.append([c["name"] + " — var. anual", c.get("yoyText"), kpi.get("ultimoP")])

    r = r0 + 1
    for row in rows:
        for j, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = TXT
            if r % 2 == 0:
                cell.fill = BAND_FILL
        r += 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
    _autow(ws, [40, 24, 20])
    wb.save(out_path)


def _build_pib_workbook(ind: dict, out_path: Path):
    """Genera el Excel individual de PIB con dos hojas: PIB oportuno y Nivel PIB."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    add_pib_sheets(wb, {"indicators": {"PIB": ind}})
    wb.save(out_path)


def _write_subset_sheet(ws, ind: dict, title: str, subtitle: str, include_indices: list[int]) -> None:
    """Escribe una hoja con un subconjunto de columnas del indicador."""
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE
    ws["A2"] = subtitle
    ws["A2"].font = MUT
    all_cols = ind.get("columns", [])
    cols = [all_cols[i] for i in include_indices if i < len(all_cols)]
    headers = ["Periodo", "Fecha"] + [c["label"] for c in cols] + [
        "Carácter del dato", "Fuente", "URL del boletín", "Fecha de publicación"
    ]
    r0 = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=r0, column=i, value=h)
    _style_header(ws, r0, len(headers))

    obs = ind.get("observations", [])
    r = r0 + 1
    for o in obs:
        period = o["period"]
        d = _period_date(period)
        ws.cell(row=r, column=1, value=period).font = TXT
        ws.cell(row=r, column=2, value=d).font = TXT
        if d:
            ws.cell(row=r, column=2).number_format = "yyyy-mm-dd"
        values = list(o.get("values", []))
        for j, idx in enumerate(include_indices, start=3):
            if idx < len(values) and j - 3 < len(cols):
                v = values[idx]
                cell = ws.cell(row=r, column=j, value=v)
                cell.font = TXT
                cell.border = BORDER
                cell.number_format = _xl_fmt(cols[j - 3].get("fmt", "num"))
                if r % 2 == 0:
                    cell.fill = BAND_FILL
        meta_start = 3 + len(cols)
        meta = [
            _caracter_dato(period),
            ind.get("fuente", {}).get("nombre", "—"),
            ind.get("url_boletin_oficial") or ind.get("fuente", {}).get("link") or "—",
            ind.get("fecha_publicacion") or "—",
        ]
        for j, v in enumerate(meta, start=meta_start):
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = TXT
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = BAND_FILL
        r += 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    _autow(ws, [16, 13] + [26] * len(cols) + [16, 28, 55, 22])


def _build_bcmm_workbook(ind: dict, out_path: Path) -> None:
    """Genera el Excel individual de BCMM con 4 hojas temáticas."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    src = ind.get("fuente", {}).get("nombre", "INEGI")
    url = ind.get("url_boletin_oficial") or ind.get("fuente", {}).get("link") or "—"
    pub = ind.get("fecha_publicacion") or "—"
    sub = f"Fuente: {src} · Frecuencia: Mensual · Unidad: millones de dólares · Boletín: {url} · Fecha de publicación: {pub}"

    ws1 = wb.create_sheet("Comercio total")
    _write_subset_sheet(ws1, ind, f"{ind.get('nombre', 'BCMM')} — Comercio total", sub, [0, 1, 2, 26, 27, 28])
    ws2 = wb.create_sheet("Variaciones anuales")
    _write_subset_sheet(ws2, ind, f"{ind.get('nombre', 'BCMM')} — Variaciones anuales", sub, [3, 4, 5, 10, 11, 12, 13, 17, 18, 19, 23, 24, 25])
    ws3 = wb.create_sheet("Petrolero y no petrolero")
    _write_subset_sheet(ws3, ind, f"{ind.get('nombre', 'BCMM')} — Petrolero y no petrolero", sub, [0, 1, 6, 7, 8, 9, 20, 21, 22])
    ws4 = wb.create_sheet("Importaciones por tipo")
    _write_subset_sheet(ws4, ind, f"{ind.get('nombre', 'BCMM')} — Importaciones por tipo de bien", sub, [1, 14, 15, 16, 17, 18, 19])
    wb.save(out_path)


def _build_individual_workbook(ind: dict, cfg: dict, kpicfg: dict, out_path: Path):
    """Genera un Excel con una sola hoja visible por indicador."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(ind["key"])
    _write_indicator_sheet(ws, ind)
    wb.save(out_path)


def _write_indicator_sheet(ws, ind: dict, sheet_title: str | None = None) -> None:
    """Escribe la hoja de datos principal de un indicador."""
    ws.sheet_view.showGridLines = False
    ws["A1"] = sheet_title or ind.get("nombre", ind["key"])
    ws["A1"].font = TITLE
    ws["A2"] = (f"Fuente: {ind.get('fuente', {}).get('nombre', '—')} · "
                f"Frecuencia: {ind.get('frecuencia', '—')} · "
                f"Unidad: {ind.get('unidad', '—')}")
    ws["A2"].font = MUT
    ws["A3"] = (f"URL del boletín / serie: {ind.get('url_boletin_oficial') or ind.get('fuente', {}).get('link') or '—'} · "
                f"Fecha de publicación: {ind.get('fecha_publicacion') or '—'}")
    ws["A3"].font = MUT

    cols = ind.get("columns", [])
    headers = ["Periodo", "Fecha"] + [c["label"] for c in cols] + [
        "Carácter del dato", "Fuente", "URL del boletín", "Fecha de publicación"
    ]
    r0 = 5
    for i, h in enumerate(headers, start=1):
        ws.cell(row=r0, column=i, value=h)
    _style_header(ws, r0, len(headers))

    obs = ind.get("observations", [])
    r = r0 + 1
    for o in obs:
        period = o["period"]
        d = _period_date(period)
        ws.cell(row=r, column=1, value=period).font = TXT
        ws.cell(row=r, column=2, value=d).font = TXT
        if d:
            ws.cell(row=r, column=2).number_format = "yyyy-mm-dd"

        values = list(o.get("values", []))
        for j, v in enumerate(values, start=3):
            if j - 3 < len(cols):
                cell = ws.cell(row=r, column=j, value=v)
                cell.font = TXT
                cell.border = BORDER
                cell.number_format = _xl_fmt(cols[j - 3].get("fmt", "num"))
                if r % 2 == 0:
                    cell.fill = BAND_FILL

        meta_start = 3 + len(cols)
        meta = [
            _caracter_dato(period),
            ind.get("fuente", {}).get("nombre", "—"),
            ind.get("url_boletin_oficial") or ind.get("fuente", {}).get("link") or "—",
            ind.get("fecha_publicacion") or "—",
        ]
        for j, v in enumerate(meta, start=meta_start):
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = TXT
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = BAND_FILL

        r += 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
    _autow(ws, [16, 13] + [22] * len(cols) + [16, 28, 55, 22])


def _write_emim_subsector_sheet(ws, ind: dict) -> None:
    """Escribe la hoja de subsectores del EMIM."""
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{ind.get('nombre', 'EMIM')} — Desempeño por subsector"
    ws["A1"].font = TITLE
    ws["A2"] = (f"Fuente: {ind.get('fuente', {}).get('nombre', '—')} · "
                f"Periodo: {ind.get('last_observation', '—')} · "
                "Nota: variaciones anuales en cifras originales.")
    ws["A2"].font = MUT

    headers = ["Código", "Subsector",
               "Producción índice", "Producción anual (%)",
               "Personal índice", "Personal anual (%)",
               "Horas índice", "Horas anual (%)",
               "Remuneraciones índice", "Remuneraciones anual (%)"]
    r0 = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=r0, column=i, value=h)
    _style_header(ws, r0, len(headers))

    detail = ind.get("subsectores_detalle") or ind.get("subsectores")
    if not detail:
        ws.cell(row=r0 + 1, column=1, value="Sin datos de subsectores disponibles.").font = MUT
        ws.merge_cells(start_row=r0 + 1, start_column=1, end_row=r0 + 1, end_column=len(headers))
        _autow(ws, [12, 55, 18, 18, 18, 18, 18, 18, 24, 24])
        return

    sample = next(iter(detail.values()), {})
    if not isinstance(sample, dict):
        ws.cell(row=r0 + 1, column=1, value="Sin detalle de subsectores disponible.").font = MUT
        ws.merge_cells(start_row=r0 + 1, start_column=1, end_row=r0 + 1, end_column=len(headers))
        _autow(ws, [12, 55, 18, 18, 18, 18, 18, 18, 24, 24])
        return

    r = r0 + 1
    for code in sorted(detail.keys(), key=lambda c: int(c) if isinstance(c, str) and c.isdigit() else str(c)):
        info = detail[code]
        if not isinstance(info, dict):
            continue
        values = [
            code,
            info.get("nombre", ""),
            info.get("produccion_index"),
            info.get("produccion_anual"),
            info.get("personal_index"),
            info.get("personal_anual"),
            info.get("horas_index"),
            info.get("horas_anual"),
            info.get("remuneraciones_index"),
            info.get("remuneraciones_anual"),
        ]
        for i, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = TXT
            cell.border = BORDER
            if i in (3, 5, 7, 9):
                cell.number_format = _xl_fmt("idx")
            elif i in (4, 6, 8, 10):
                cell.number_format = _xl_fmt("pct-frac")
            if r % 2 == 0:
                cell.fill = BAND_FILL
        r += 1

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    _autow(ws, [12, 55, 18, 18, 18, 18, 18, 18, 24, 24])


def build_emim_workbook(ind: dict, out_path: Path) -> None:
    """Genera el Excel individual de EMIM con dos hojas: Indicadores y Subsectores."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet("Indicadores")
    _write_indicator_sheet(ws1, ind, sheet_title=ind.get("nombre", "EMIM"))
    ws2 = wb.create_sheet("Subsectores")
    _write_emim_subsector_sheet(ws2, ind)
    wb.save(out_path)


def build_individual_files(payload: dict, pilot: list[str] | None = None):
    """Genera un archivo Excel por indicador y actualiza flags en el payload.

    Si `pilot` se provee, solo genera archivos para esas claves y marca los
    demás como no disponibles para no dejar enlaces rotos en el repositorio.
    """
    kpicfg = get_cfg("KPICFG")
    IND_DIR.mkdir(parents=True, exist_ok=True)
    pilot_set = set(pilot) if pilot else None
    for key, ind in payload["indicators"].items():
        cfg = kpicfg.get(key)
        out_dir = IND_DIR / key
        out_path = out_dir / f"{key}_datos.xlsx"
        if pilot_set and key not in pilot_set:
            ind["xlsx_disponible"] = False
            ind["xlsx_causa"] = "Producto en fase piloto; próximamente disponible"
            ind["url_excel_individual"] = None
            if out_path.exists():
                out_path.unlink()
            continue
        out_dir.mkdir(parents=True, exist_ok=True)
        if not ind.get("observations") or not cfg:
            ind["xlsx_disponible"] = False
            ind["xlsx_causa"] = "Sin observaciones o sin configuración de métricas"
            ind["url_excel_individual"] = None
            continue
        try:
            if key == "PIB":
                _build_pib_workbook(ind, out_path)
            elif key == "PIBSEC":
                _build_pibsec_workbook(ind, out_path)
            elif key == "EMIM" and len(ind.get("columns", [])) >= 18:
                build_emim_workbook(ind, out_path)
            elif key == "IMAI" and len(ind.get("columns", [])) >= 18:
                _build_imai_workbook(ind, out_path)
            elif key == "BCMM" and len(ind.get("columns", [])) >= 29:
                _build_bcmm_workbook(ind, out_path)
            else:
                _build_individual_workbook(ind, cfg, kpicfg, out_path)
            ind["xlsx_disponible"] = True
            ind["url_excel_individual"] = str(out_path.relative_to(ROOT))
            ind["xlsx_causa"] = None
        except Exception as e:
            ind["xlsx_disponible"] = False
            ind["xlsx_causa"] = f"Error al generar Excel: {e}"
            ind["url_excel_individual"] = None


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = H
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER


def _autow(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_resumen(wb, payload, manifest):
    ws = wb.create_sheet("Síntesis de coyuntura", 0)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Indicadores macroeconómicos de México — Síntesis de coyuntura"
    ws["A1"].font = TITLE
    ws["A2"] = f"Generado: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')}"
    ws["A2"].font = MUT
    headers = ["Indicador", "Fuente", "Frecuencia", "Última observación", "Último valor", "Variación", "Unidad"]
    r0 = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=r0, column=i, value=h)
    _style_header(ws, r0, len(headers))
    order = payload.get("order") or list(payload["indicators"].keys())
    r = r0 + 1
    for key in order:
        ind = payload["indicators"].get(key)
        if not ind:
            continue
        kpi = (ind.get("metrics") or {}).get("kpi")
        if kpi:
            periodo = kpi.get("ultimoP") or ind.get("last_observation")
            val = kpi.get("ultimoFmt")
            var = kpi.get("varText")
        else:
            obs = ind["observations"]
            last = obs[-1] if obs else {"values": [None]}
            val = last["values"][0] if last["values"] else None
            var = last["values"][1] if len(last.get("values", [])) > 1 else None
            periodo = ind.get("last_observation")
        row_vals = [ind.get("nombre"), ind.get("fuente", {}).get("nombre"), ind.get("frecuencia"),
                    periodo, val, var, ind.get("unidad")]
        for i, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = TXT
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = BAND_FILL
        r += 1
    ws["A" + str(r + 1)] = ("Nota: cifras oficiales sujetas a revisión. Análisis por reglas deterministas; "
                            "no se atribuye causalidad. La fuente de la tasa de desocupación es INEGI/ENOE. "
                            "Los indicadores en estado 'pendiente de token' o 'dato de respaldo' no se "
                            "presentan como actualización automática (ver 'Control de actualizaciones').")
    ws["A" + str(r + 1)].font = MUT
    ws["A" + str(r + 1)].alignment = WRAP
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 2, end_column=7)
    _autow(ws, [42, 30, 14, 20, 16, 14, 30])


def add_metodologia(wb, payload):
    ws = wb.create_sheet("Metodología y fuentes")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Metodología y fuentes"
    ws["A1"].font = TITLE
    headers = ["Indicador", "Fuente oficial", "Serie", "Unidad", "Ajuste estacional",
               "Frecuencia", "Método de actualización", "Enlace"]
    r0 = 3
    for i, h in enumerate(headers, start=1):
        ws.cell(row=r0, column=i, value=h)
    _style_header(ws, r0, len(headers))
    order = payload.get("order") or list(payload["indicators"].keys())
    r = r0 + 1
    for key in order:
        ind = payload["indicators"].get(key)
        if not ind:
            continue
        f = ind.get("fuente", {})
        vals = [ind.get("nombre"), f.get("nombre"), f.get("serie") or "—", ind.get("unidad"),
                ind.get("ajuste_estacional"), ind.get("frecuencia"), f.get("metodo"), f.get("link")]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = TXT
            cell.border = BORDER
            cell.alignment = WRAP
            if r % 2 == 0:
                cell.fill = BAND_FILL
        r += 1
    _autow(ws, [40, 34, 12, 26, 18, 14, 26, 40])


def _next_pub_map(calendar):
    """clave -> primera publicación próxima, no anunciada o evento futuro."""
    out = {}
    def _sort_key(x):
        iso = x.get("fecha_iso")
        return (iso or "", 0 if iso else 1)
    items = sorted((calendar or {}).get("items", []), key=_sort_key)
    for it in items:
        if it.get("estatus") in ("próximo", "no_anunciada", "evento") and it.get("clave") not in out:
            out[it["clave"]] = it
    return out


def add_control(wb, payload, manifest, log, calendar=None):
    ws = wb.create_sheet("Control de actualizaciones")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Control de actualizaciones"
    ws["A1"].font = TITLE
    ws["A2"] = (f"Última corrida del pipeline: {log.get('finished_at', '—')} · "
                f"resultado: {log.get('result', '—')} · modo: {log.get('mode', '—')}")
    ws["A2"].font = MUT
    headers = ["Indicador", "Clasificación", "Estado", "Origen del dato", "Requiere token",
               "Serie confirmada", "Última observación", "Periodo de referencia",
               "Fecha de publicación", "Próxima publicación (calendario)", "Fecha de consulta",
               "Actualización archivo", "Revisión detectada", "Observaciones de calidad"]
    nextpub = _next_pub_map(calendar)
    r0 = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=r0, column=i, value=h)
    _style_header(ws, r0, len(headers))
    rows = {m["clave"]: m for m in manifest.get("indicadores", [])}
    order = payload.get("order") or list(payload["indicators"].keys())
    r = r0 + 1
    for key in order:
        m = rows.get(key)
        if not m:
            continue
        np = nextpub.get(key)
        np_txt = f"{np['fecha_publicacion']} · {np['periodo_referencia']}" if np else "—"
        vals = [m["indicador"], m.get("clasificacion"), m.get("estado"), m.get("origen_dato"),
                m.get("requiere_token") or "—", "Sí" if m.get("serie_confirmada") else "No",
                m["ultima_observacion"], m.get("periodo_referencia"), m.get("fecha_publicacion"),
                np_txt, m["fecha_consulta"], m["fecha_actualizacion_archivo"],
                "Sí" if m["revision_detectada"] else "No", m["observaciones"]]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = TXT
            cell.border = BORDER
            cell.alignment = WRAP
            if m["revision_detectada"]:
                cell.fill = PatternFill("solid", fgColor="FFF6E2E8")
            elif r % 2 == 0:
                cell.fill = BAND_FILL
        r += 1
    # Advertencias del pipeline
    r += 2
    ws.cell(row=r, column=1, value="Advertencias del pipeline:").font = LBL
    for w in log.get("warnings", []):
        r += 1
        ws.cell(row=r, column=1, value="• " + w).font = MUT
    _autow(ws, [40, 15, 26, 14, 14, 15, 18, 18, 26, 28, 16, 18, 16, 50])


# Hojas de datos a crear (o recrear) para indicadores principales.
NEW_SHEETS = {
    "PIBSEC": "PIB Sectorial",
    "IMFBCF": "Formación bruta capital fijo",
    "IOAE": "IOAE",
    "EMIM": "EMIM (Manufactura)",
    "BCMM": "Balanza comercial",
    "IMAI": "IMAI",
}

# Hojas heredadas del libro base que ya no corresponden al perfil V3.
LEGACY_SHEETS = ["Exportaciones", "PIB"]


def add_indicator_sheets(wb, payload):
    """Crea o recrea hojas de datos para los indicadores principales. Si aún no
    tienen observaciones (scaffold pendiente de token), la hoja queda con los
    encabezados y una nota honesta, sin cifras inventadas."""
    for key, sheet_name in NEW_SHEETS.items():
        ind = payload["indicators"].get(key)
        if not ind:
            continue
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws["A1"] = ind.get("nombre")
        ws["A1"].font = TITLE
        ws["A2"] = (f"Fuente: {ind.get('fuente', {}).get('nombre', '—')} · "
                    f"Frecuencia: {ind.get('frecuencia', '—')} · Unidad: {ind.get('unidad', '—')} · "
                    f"Estado: {ind.get('estado', '—')}")
        ws["A2"].font = MUT
        cols = ind.get("columns", [])
        headers = ["Periodo"] + [c["label"] for c in cols]
        r0 = 4
        for i, h in enumerate(headers, start=1):
            ws.cell(row=r0, column=i, value=h)
        _style_header(ws, r0, len(headers))
        obs = ind.get("observations", []) or []
        if not obs:
            note = ("Sin observaciones cargadas todavía. "
                    f"Se activará al configurar {ind.get('requiere_token', 'el token')}_TOKEN "
                    "y confirmar la serie oficial. No se muestran cifras estimadas ni inventadas.")
            ws.cell(row=r0 + 1, column=1, value=note).font = MUT
            ws.merge_cells(start_row=r0 + 1, start_column=1, end_row=r0 + 2, end_column=max(2, len(headers)))
            ws.cell(row=r0 + 1, column=1).alignment = WRAP
        else:
            r = r0 + 1
            for o in obs:
                ws.cell(row=r, column=1, value=o["period"]).font = TXT
                for i, v in enumerate(o["values"], start=2):
                    cell = ws.cell(row=r, column=i, value=v)
                    cell.font = TXT
                    if i - 2 < len(cols):
                        cell.number_format = _xl_fmt(cols[i - 2].get("fmt", "num"))
                r += 1
        _autow(ws, [16] + [22] * len(cols))

        if key == "EMIM" and len(ind.get("columns", [])) >= 18:
            sub_name = "Subsectores EMIM"
            if sub_name in wb.sheetnames:
                wb.remove(wb[sub_name])
            ws2 = wb.create_sheet(sub_name)
            _write_emim_subsector_sheet(ws2, ind)


def add_pib_sheets(wb, payload):
    """Crea las hojas separadas PIB oportuno (EOPIBT) y Nivel PIB (PIBT)."""
    ind = payload["indicators"].get("PIB")
    if not ind:
        return
    for name in ("PIB oportuno", "Nivel PIB"):
        if name in wb.sheetnames:
            wb.remove(wb[name])

    # Hoja 1: PIB oportuno (variaciones)
    ws = wb.create_sheet("PIB oportuno")
    ws.sheet_view.showGridLines = False
    ws["A1"] = ind.get("nombre")
    ws["A1"].font = TITLE
    ws["A2"] = (f"Fuente: {ind.get('fuente', {}).get('nombre', '—')} · "
                f"Frecuencia: {ind.get('frecuencia', '—')} · Unidad: {ind.get('unidad', '—')} · "
                f"Estado: {ind.get('estado', '—')}")
    ws["A2"].font = MUT

    cols = [
        {"label": "Var. trimestral desest. (%)", "fmt": "pct-frac"},
        {"label": "Var. anual desest. (%)", "fmt": "pct-frac"},
        {"label": "Var. anual original (%)", "fmt": "pct-frac"},
        {"label": "Acumulado (%)", "fmt": "pct-frac"},
    ]
    headers = ["Periodo", "Fecha"] + [c["label"] for c in cols] + ["Carácter", "Fuente", "URL boletín", "Fecha publicación"]
    r0 = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=r0, column=i, value=h)
    _style_header(ws, r0, len(headers))

    obs = ind.get("observations", [])
    r = r0 + 1
    for o in obs:
        d = _period_date(o["period"])
        ws.cell(row=r, column=1, value=o["period"]).font = TXT
        ws.cell(row=r, column=2, value=d).font = TXT
        if d:
            ws.cell(row=r, column=2).number_format = "yyyy-mm-dd"
        for i, v in enumerate(o["values"], start=3):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = TXT
            cell.number_format = _xl_fmt(cols[i - 3].get("fmt", "num"))
        ws.cell(row=r, column=7, value=_caracter_dato(o["period"])).font = TXT
        ws.cell(row=r, column=8, value=ind.get("fuente", {}).get("nombre", "—")).font = TXT
        ws.cell(row=r, column=9, value=ind.get("url_boletin_oficial") or ind.get("fuente", {}).get("link") or "—").font = TXT
        ws.cell(row=r, column=10, value=ind.get("fecha_publicacion") or "—").font = TXT
        if r % 2 == 0:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = BAND_FILL
        r += 1
    _autow(ws, [16, 13, 24, 22, 24, 18, 14, 28, 55, 22])

    # Hoja 2: Nivel PIB
    pibt = ind.get("pibt")
    ws2 = wb.create_sheet("Nivel PIB")
    ws2.sheet_view.showGridLines = False
    if pibt:
        ws2["A1"] = "PIB trimestral a precios constantes de 2018"
        ws2["A1"].font = TITLE
        ws2["A2"] = (f"Fuente: {pibt.get('fuente', {}).get('nombre', '—')} · "
                     f"Serie: {pibt.get('fuente', {}).get('serie', '—')} · "
                     f"Frecuencia: {pibt.get('frecuencia', '—')} · "
                     f"Unidad: {pibt.get('unidad', '—')}")
        ws2["A2"].font = MUT
        pibt_cols = ["Periodo", "Fecha", "Nivel", "Unidad", "Carácter", "Fuente", "Serie"]
        r0 = 4
        for i, h in enumerate(pibt_cols, start=1):
            ws2.cell(row=r0, column=i, value=h)
        _style_header(ws2, r0, len(pibt_cols))
        pibt_obs = pibt.get("observations", [])
        r = r0 + 1
        for o in pibt_obs:
            d = _period_date(o["period"])
            ws2.cell(row=r, column=1, value=o["period"]).font = TXT
            ws2.cell(row=r, column=2, value=d).font = TXT
            if d:
                ws2.cell(row=r, column=2).number_format = "yyyy-mm-dd"
            ws2.cell(row=r, column=3, value=o["values"][0]).number_format = _xl_fmt("bill")
            ws2.cell(row=r, column=4, value=pibt.get("unidad", "—")).font = TXT
            ws2.cell(row=r, column=5, value=_caracter_dato(o["period"])).font = TXT
            ws2.cell(row=r, column=6, value=pibt.get("fuente", {}).get("nombre", "—")).font = TXT
            ws2.cell(row=r, column=7, value=pibt.get("fuente", {}).get("serie", "—")).font = TXT
            if r % 2 == 0:
                for c in range(1, len(pibt_cols) + 1):
                    ws2.cell(row=r, column=c).fill = BAND_FILL
            r += 1
        _autow(ws2, [16, 13, 18, 40, 14, 18, 18])
    else:
        ws2["A1"] = "Nivel PIB"
        ws2["A1"].font = TITLE
        ws2["A2"] = "Sin datos de nivel PIBT disponibles todavía."
        ws2["A2"].font = MUT


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pilot", action="store_true", help="Genera archivos individuales solo para IGAE e INPC")
    # Al ser importado por pytest, evita que los argumentos de pytest (-q) lleguen aquí.
    if "pytest" in sys.modules:
        args = ap.parse_args([])
    else:
        args = ap.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    payload = L.load_data()
    manifest = json.loads((L.DATA_DIR / "manifest.json").read_text(encoding="utf-8"))
    log_path = L.DATA_DIR / "update_log.json"
    log = json.loads(log_path.read_text(encoding="utf-8")) if log_path.exists() else {}

    if BASE_XLSX.exists():
        wb = openpyxl.load_workbook(BASE_XLSX)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    cal_path = L.DATA_DIR / "calendario_publicaciones.json"
    calendar = json.loads(cal_path.read_text(encoding="utf-8")) if cal_path.exists() else {}

    # No alterar las hojas originales, salvo las que se recrean o quedaron obsoletas.
    for name in ("Resumen ejecutivo", "Síntesis de coyuntura", "Metodología y fuentes", "Control de actualizaciones"):
        if name in wb.sheetnames:
            wb.remove(wb[name])
    for name in NEW_SHEETS.values():
        if name in wb.sheetnames:
            wb.remove(wb[name])
    for name in LEGACY_SHEETS:
        if name in wb.sheetnames:
            wb.remove(wb[name])
    if "Subsectores EMIM" in wb.sheetnames:
        wb.remove(wb["Subsectores EMIM"])
    add_pib_sheets(wb, payload)
    add_indicator_sheets(wb, payload)
    add_metodologia(wb, payload)
    add_control(wb, payload, manifest, log, calendar)
    add_resumen(wb, payload, manifest)  # queda como primera hoja (índice 0)

    wb.save(OUT_XLSX)
    print(f"OK: {OUT_XLSX.relative_to(ROOT)} ({OUT_XLSX.stat().st_size} bytes) · hojas: {wb.sheetnames}")

    # Archivos individuales por indicador (actualiza flags en indicadores.json).
    build_individual_files(payload, pilot=PILOT_KEYS if args.pilot else None)
    (L.DATA_DIR / "indicadores.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
    )
    print(f"OK: archivos individuales generados en {IND_DIR.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
