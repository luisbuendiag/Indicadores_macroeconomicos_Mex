"""Conector de Inversión Extranjera Directa (IED) — Secretaría de Economía.

Fuente primaria:
  - Resumen histórico (cifras originalmente publicadas y actualizadas):
    https://www.gob.mx/cms/uploads/attachment/file/1100772/Datos_originales_y_actualizacion__1_.xlsx
  - Desglose por tipo, entidad federativa, país de origen y actividad económica:
    https://www.gob.mx/cms/uploads/attachment/file/1100604/2026_2T_Flujos_TI_AC_3.xlsx

El conector descarga los archivos de gob.mx, los parsea y construye el
indicador IED con dos series conceptuales:

  - observations: flujo trimestral (IED del trimestre exclusivo), con
    componentes por tipo de inversión.
  - metrics.acumulado: serie acumulada al corte (ENE-MAR, ENE-JUN, ...).

Para el 2T-2026 se combina con data/ied_manual_2026_2t.json mientras la
Secretaría de Economía actualice el XLSX estructurado.
"""
from __future__ import annotations

import json
import os
import re
import urllib.request
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from .base import SourceResult, USER_AGENT

ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT / "data"
DOWNLOADS_DIR = DATA_DIR / "downloads" / "ied"

URL_RESUMEN = "https://www.gob.mx/cms/uploads/attachment/file/1100772/Datos_originales_y_actualizacion__1_.xlsx"
URL_FlUJOS_TI_AC = "https://www.gob.mx/cms/uploads/attachment/file/1100604/2026_2T_Flujos_TI_AC_3.xlsx"

MESES_ABR = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


def _clean(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).replace(",", "").strip()
        if s in ("-", "", "C", "R", "ND", "X"):
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


def _period_to_quarter(periodo: str) -> tuple[int, int] | None:
    """'Enero - marzo' -> (1,3); 'Enero - junio' -> (4,6) etc."""
    p = (periodo or "").lower().replace(" de ", " ").strip()
    p = re.sub(r"\s+", " ", p)
    # normaliza guiones sin espacio
    p = re.sub(r"enero-", "enero - ", p)
    p = re.sub(r"marzo-", "marzo - ", p)
    p = re.sub(r"junio-", "junio - ", p)
    p = re.sub(r"septiembre-", "septiembre - ", p)
    p = re.sub(r"diciembre-", "diciembre - ", p)

    mapping = {
        "enero - marzo": (1, 3),
        "enero marzo": (1, 3),
        "enero - junio": (4, 6),
        "enero junio": (4, 6),
        "enero - septiembre": (7, 9),
        "enero septiembre": (7, 9),
        "enero - diciembre": (10, 12),
        "enero diciembre": (10, 12),
    }
    return mapping.get(p)


def _quarter_label(year: int, end_month: int) -> str:
    q = (end_month - 1) // 3 + 1
    return f"{q}T-{year % 100:02d}"


def _accum_label(year: int, end_month: int) -> str:
    start = f"{MESES_ABR[0]}"
    end = f"{MESES_ABR[end_month - 1]}"
    return f"{start}-{end} {year % 100:02d}"


def _download(url: str, dest: Path, timeout: int = 60) -> bool:
    dest.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            dest.write_bytes(r.read())
        return True
    except Exception as e:  # noqa: BLE001
        return False


def _load_manual_2t() -> dict | None:
    path = DATA_DIR / "ied_manual_2026_2t.json"
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def _parse_resumen(path: Path, warnings: list[str]) -> list[dict]:
    """Parsea la hoja de resumen histórico y devuelve observaciones acumuladas."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    out: list[dict] = []
    for r in range(4, ws.max_row + 1):
        anio = ws.cell(r, 1).value
        periodo = ws.cell(r, 2).value
        publicado = _clean(ws.cell(r, 3).value)
        actualizado = _clean(ws.cell(r, 4).value)
        if not anio or not periodo or not isinstance(anio, (int, float)):
            continue
        anio = int(anio)
        qmap = _period_to_quarter(str(periodo))
        if not qmap:
            continue
        _, end_month = qmap
        # Preferir dato actualizado; si es '-', usar publicado.
        valor = actualizado if actualizado is not None else publicado
        if valor is None:
            continue
        out.append({
            "year": anio,
            "end_month": end_month,
            "period_acum": _accum_label(anio, end_month),
            "quarter": _quarter_label(anio, end_month),
            "valor_publicado": publicado,
            "valor_actualizado": actualizado,
            "acumulado": valor,
        })
    return out


def _parse_flujos_ti(path: Path, warnings: list[str]) -> dict | None:
    """Parsea el archivo de flujos por tipo de inversión.

    Devuelve el total y componentes para el último trimestre disponible (2026 1T).
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:  # noqa: BLE001
        warnings.append(f"ied: no se pudo abrir flujos TI: {e}")
        return None
    ws = wb["Por Tipo de Inversión"]
    mc = ws.max_column
    # header: row 3 years, row 4 quarters; data starts row 5
    # take last 5 columns to find 2026 Q1 and any Q2
    years = [ws.cell(3, c).value for c in range(mc - 4, mc + 1)]
    qtrs = [ws.cell(4, c).value for c in range(mc - 4, mc + 1)]
    vals = [ws.cell(5, c).value for c in range(mc - 4, mc + 1)]
    # find 2026 Q1
    idx = None
    for i, (y, q) in enumerate(zip(years, qtrs)):
        if y == 2026 and q == 1:
            idx = i
            break
    if idx is None:
        warnings.append("ied: no se encontró 2026 1T en flujos TI")
        return None
    col = (mc - 4) + idx

    def _row_value(label: str) -> float | None:
        for r in range(6, min(ws.max_row, 30) + 1):
            if ws.cell(r, 1).value == label:
                return _clean(ws.cell(r, col).value)
        return None

    return {
        "year": 2026,
        "quarter": 1,
        "total": _clean(vals[idx]),
        "nuevas": _row_value("Nuevas inversiones"),
        "reinversion": _row_value("Reinversión de utilidades"),
        "cuentas": _row_value("Cuentas entre compañías"),
    }


def _historical_flows(acum: list[dict]) -> list[dict]:
    """Convierte serie acumulada en flujos trimestrales.

    Ordena por año y end_month, calcula diferencias entre acumulados
    consecutivos dentro del mismo año.
    """
    acum_sorted = sorted(acum, key=lambda x: (x["year"], x["end_month"]))
    flows: list[dict] = []
    prev_year = None
    prev_end = None
    prev_val = None
    for a in acum_sorted:
        # si es 1T del año (enero-marzo), el flujo = acumulado
        if a["end_month"] == 3:
            flujo = a["acumulado"]
            prev_year = a["year"]
            prev_end = a["end_month"]
            prev_val = a["acumulado"]
        else:
            if a["year"] == prev_year and prev_val is not None:
                flujo = a["acumulado"] - prev_val
                prev_val = a["acumulado"]
                prev_end = a["end_month"]
            else:
                flujo = None
        flows.append({
            "year": a["year"],
            "end_month": a["end_month"],
            "period_acum": a["period_acum"],
            "quarter": a["quarter"],
            "acumulado": a["acumulado"],
            "flujo": flujo,
        })
    return flows


def _pct_to_frac(items: list[dict] | None) -> list[dict]:
    """Convierte participaciones de porcentaje (88.5) a fracción (0.885)."""
    if not items:
        return []
    out = []
    for it in items:
        it2 = dict(it)
        for k in ("participacion_pct",):
            v = it2.get(k)
            if v is not None and v > 1:
                it2[k] = round(v / 100, 6)
        out.append(it2)
    return out


def _componentes_flujo_from_manual(manual: dict, ti_1t: dict | None = None) -> dict:
    """Componentes del flujo 2T calculados como diferencia entre acumulados 2T y 1T.

    Si el manual incluye componentes del 1T (clave `tipo_de_inversion_1t`), se usan.
    En otro caso se usan los componentes 1T del XLS oficial (`ti_1t`) y se ajustan
    proporcionalmente para que sumen el total 1T del manual si éste es más preciso.
    """
    comp_2t = {c["concepto"]: c["mdd"] for c in manual["tipo_de_inversion"]}
    comp_1t = None
    if manual.get("tipo_de_inversion_1t"):
        comp_1t = {c["concepto"]: c["mdd"] for c in manual["tipo_de_inversion_1t"]}
    elif ti_1t:
        comp_1t = {
            "Nuevas inversiones": ti_1t.get("nuevas"),
            "Reinversión de utilidades": ti_1t.get("reinversion"),
            "Cuentas entre compañías": ti_1t.get("cuentas"),
        }
        # Ajustar proporcionalmente al total 1T del manual
        total_xls = (comp_1t["Nuevas inversiones"] or 0) + (comp_1t["Reinversión de utilidades"] or 0) + (comp_1t["Cuentas entre compañías"] or 0)
        total_1t = manual.get("acumulado_1t", total_xls)
        if total_xls and total_xls:
            factor = total_1t / total_xls
            comp_1t = {k: round(v * factor, 2) if v is not None else None for k, v in comp_1t.items()}
    if not comp_1t:
        return {k: None for k in comp_2t}
    return {
        "Nuevas inversiones": round(comp_2t.get("Nuevas inversiones", 0) - (comp_1t.get("Nuevas inversiones") or 0), 2),
        "Reinversión de utilidades": round(comp_2t.get("Reinversión de utilidades", 0) - (comp_1t.get("Reinversión de utilidades") or 0), 2),
        "Cuentas entre compañías": round(comp_2t.get("Cuentas entre compañías", 0) - (comp_1t.get("Cuentas entre compañías") or 0), 2),
    }


def _build_indicator(resumen: list[dict], manual: dict | None, ti_1t: dict | None, warnings: list[str]) -> dict:
    """Construye el objeto indicador IED."""
    # Serie acumulada histórica
    acum = _parse_resumen(Path(resumen) if isinstance(resumen, (str, Path)) else Path(resumen), warnings)

    # Si hay manual 2T, inyectar/actualizar observación 2T-2026
    if manual:
        # buscar si ya existe 2026 2T en acum (debería estar con valor publicado)
        found = [i for i, a in enumerate(acum) if a["year"] == 2026 and a["end_month"] == 6]
        if found:
            i = found[0]
            # usar el manual actualizado
            acum[i]["acumulado"] = manual["acumulado"]
            acum[i]["valor_actualizado"] = manual["acumulado"]
        else:
            acum.append({
                "year": 2026,
                "end_month": 6,
                "period_acum": _accum_label(2026, 6),
                "quarter": _quarter_label(2026, 6),
                "valor_publicado": manual["acumulado"],
                "valor_actualizado": manual["acumulado"],
                "acumulado": manual["acumulado"],
            })

    # Calcular flujos trimestrales
    flows = _historical_flows(acum)
    flow_by_ym = {(f["year"], f["end_month"]): f["flujo"] for f in flows}
    for a in acum:
        a["flujo"] = flow_by_ym.get((a["year"], a["end_month"]))
    acum_sorted = sorted(acum, key=lambda x: (x["year"], x["end_month"]))

    # Último periodo
    last = acum_sorted[-1]

    # Componentes del flujo del último periodo.
    # Si tenemos 2T manual, calculamos flujo de componentes manualmente.
    comp_flujo: dict[str, float | None] = {"Nuevas inversiones": None, "Reinversión de utilidades": None, "Cuentas entre compañías": None}
    if manual and manual.get("tipo_de_inversion"):
        comp_flujo = _componentes_flujo_from_manual(manual, ti_1t)
    elif ti_1t:
        # solo tenemos 1T, el flujo = acumulado 1T
        comp_flujo = {
            "Nuevas inversiones": ti_1t.get("nuevas"),
            "Reinversión de utilidades": ti_1t.get("reinversion"),
            "Cuentas entre compañías": ti_1t.get("cuentas"),
        }

    # Componentes acumulados del último periodo
    comp_acum: dict[str, float | None] = {}
    if manual and manual.get("tipo_de_inversion"):
        comp_acum = {c["concepto"]: c["mdd"] for c in manual["tipo_de_inversion"]}
    elif ti_1t:
        comp_acum = {
            "Nuevas inversiones": ti_1t.get("nuevas"),
            "Reinversión de utilidades": ti_1t.get("reinversion"),
            "Cuentas entre compañías": ti_1t.get("cuentas"),
        }

    # Variación anual acumulada: preferir la del manual (fracción) o calcular
    # contra el mismo corte del año anterior (publicado original).
    last_year, last_end = last["year"], last["end_month"]
    var_acum_pct = None
    if manual and manual.get("variacion_anual_acumulada_pct") is not None:
        var_acum_pct = float(manual["variacion_anual_acumulada_pct"])
    else:
        prev = [a for a in acum if a["year"] == last_year - 1 and a["end_month"] == last_end]
        if prev:
            prev_val = prev[0].get("valor_publicado") or prev[0]["acumulado"]
            if prev_val:
                var_acum_pct = round((last["acumulado"] / prev_val - 1), 6)

    # Variación anual del flujo trimestral: preferir la del manual o calcular
    # contra el mismo trimestre del año anterior.
    var_flujo_pct = None
    if manual and manual.get("variacion_anual_trimestral_pct") is not None:
        var_flujo_pct = float(manual["variacion_anual_trimestral_pct"])
    else:
        prev_flow = [f for f in flows if f["year"] == last_year - 1 and f["end_month"] == last_end]
        if prev_flow and prev_flow[0].get("flujo") and last["flujo"]:
            var_flujo_pct = round((last["flujo"] / prev_flow[0]["flujo"] - 1), 6)

    # Observaciones: acumulados desde el inicio del año, 5 columnas.
    # [0] IED acumulada, [1] Nuevas acumuladas, [2] Reinversión acumulada,
    # [3] Cuentas entre compañías acumuladas, [4] Var. anual acumulada (fracción).
    observations = []
    comp_keys = ["Nuevas inversiones", "Reinversión de utilidades", "Cuentas entre compañías"]
    # Mapa de acumulados publicados originales para variaciones
    publicado_by_ym = {(a["year"], a["end_month"]): a["valor_publicado"] for a in acum}
    for a in acum_sorted:
        # Variación anual del acumulado (publicado original vs mismo corte año anterior)
        prev_pub = publicado_by_ym.get((a["year"] - 1, a["end_month"]))
        var = round((a["acumulado"] / prev_pub - 1), 6) if prev_pub else None
        # Para 2026 2T, preferir la variación del manual
        if a["year"] == 2026 and a["end_month"] == 6 and manual and manual.get("variacion_anual_acumulada_pct") is not None:
            var = float(manual["variacion_anual_acumulada_pct"])
        vals = [a["acumulado"], None, None, None, var]
        if a["year"] == 2026 and a["end_month"] == 6 and comp_acum:
            vals = [a["acumulado"]] + [comp_acum.get(k) for k in comp_keys] + [var]
        elif a["year"] == 2026 and a["end_month"] == 3 and ti_1t:
            vals = [a["acumulado"], ti_1t.get("nuevas"), ti_1t.get("reinversion"), ti_1t.get("cuentas"), var]
        else:
            # histórico sin desglose de componentes
            vals = [a["acumulado"], None, None, None, var]
        observations.append({
            "period": a["quarter"],
            "period_acumulado": a["period_acum"],
            "values": [round(v, 2) if v is not None else None for v in vals],
            "flujo_trimestral": round(a.get("flujo"), 2) if a.get("flujo") is not None else None,
        })

    # KPIs / métricas
    metrics = {
        "acumulado": {
            "periodo": last["period_acum"],
            "valor": round(last["acumulado"], 2),
            "unidad": "mdd",
            "variacion_anual_pct": var_acum_pct,
        },
        "flujo_trimestral": {
            "periodo": f"{last['quarter']} flujo",
            "valor": round(last["flujo"], 2) if last["flujo"] else None,
            "unidad": "mdd",
            "variacion_anual_pct": var_flujo_pct,
        },
        "composicion_tipo": _pct_to_frac(manual.get("tipo_de_inversion")) if manual else [],
        "composicion_pais": _pct_to_frac(manual.get("pais_origen")) if manual else [],
        "composicion_sector": _pct_to_frac(manual.get("sector_economico")) if manual else [],
        "composicion_entidad": _pct_to_frac(manual.get("entidad_federativa")) if manual else [],
        "componentes_acumulado": {k: comp_acum.get(k) for k in comp_keys},
    }

    return {
        "key": "IED",
        "nombre": "Inversión Extranjera Directa (IED)",
        "descripcion": "Flujo trimestral y acumulado anual de inversión extranjera directa en México. "
                       "El valor principal muestra el acumulado ene-jun 2026. "
                       "Incluye desgloses por tipo de inversión, país de origen, sector económico y entidad federativa.",
        "frecuencia": "Trimestral",
        "unidad": "Millones de dólares",
        "ajuste_estacional": "No aplica",
        "grupo": "inversion",
        "clasificacion": "complementario",
        "columns": [
            {"label": "IED acumulada", "index": 0, "fmt": "usd"},
            {"label": "Nuevas inversiones", "index": 1, "fmt": "usd"},
            {"label": "Reinversión de utilidades", "index": 2, "fmt": "usd"},
            {"label": "Cuentas entre compañías", "index": 3, "fmt": "usd"},
            {"label": "Var. anual acumulada", "index": 4, "fmt": "pct-frac"},
        ],
        "observations": observations,
        "last_observation": last["quarter"],
        "periodo_referencia": last["period_acum"],
        "url_boletin_oficial": (manual or {}).get("url_boletin", URL_RESUMEN),
        "fuente": {
            "nombre": "Secretaría de Economía — Dirección General de Inversión Extranjera / RNIE",
            "metodo": "Descarga y parseo de archivos XLSX oficiales",
            "link": "https://www.gob.mx/se/acciones-y-programas/competitividad-y-normatividad-inversion-extranjera-directa",
            "serie": None,
        },
        "metrics": metrics,
        "api_meta": {
            "n_obs": len(observations),
            "ultima_observacion": last["quarter"],
            "ultimo_valor": last["acumulado"],
        },
    }


def fetch(config: dict | None = None) -> SourceResult:
    """Consulta los XLSX oficiales de la Secretaría de Economía y construye IED."""
    warnings: list[str] = []
    DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)

    resumen_path = DOWNLOADS_DIR / "Datos_originales_y_actualizacion.xlsx"
    flujos_path = DOWNLOADS_DIR / "2026_2T_Flujos_TI_AC.xlsx"

    ok = _download(URL_RESUMEN, resumen_path)
    if not ok:
        warnings.append("ied: no se pudo descargar el resumen histórico")
    _download(URL_FlUJOS_TI_AC, flujos_path)

    manual = _load_manual_2t()
    if not manual:
        warnings.append("ied: no existe data/ied_manual_2026_2t.json")

    ti_1t = None
    if flujos_path.exists():
        ti_1t = _parse_flujos_ti(flujos_path, warnings)

    if not resumen_path.exists() and not manual:
        return SourceResult(False, warnings=warnings)

    ind = _build_indicator(resumen_path, manual, ti_1t, warnings)
    return SourceResult(True, data={"IED": ind}, warnings=warnings)


if __name__ == "__main__":
    r = fetch()
    print("ok", r.ok)
    print("warnings", r.warnings)
    import json
    print(json.dumps(r.data.get("IED", {}).get("metrics"), indent=2, ensure_ascii=False))
